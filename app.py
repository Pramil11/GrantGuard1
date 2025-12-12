from flask import Flask, render_template, request, redirect, session, url_for, make_response, send_file
import psycopg2
from psycopg2 import errors as psycopg2_errors
from psycopg2.extras import RealDictCursor
import os
import json
from datetime import date
from io import BytesIO
from urllib.parse import quote
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from openai import OpenAI

from dotenv import load_dotenv
load_dotenv()

app = Flask(__name__, template_folder='Templates')
app.secret_key = "change-this-to-any-random-secret"  # needed for session

# Add custom Jinja2 filter for JSON parsing
@app.template_filter('from_json')
def from_json_filter(value):
    """Parse JSON string in templates."""
    if not value:
        return None
    try:
        if isinstance(value, str):
            return json.loads(value)
        return value
    except (json.JSONDecodeError, TypeError):
        return None

# ---- Admin global budget (1 million) ----
ADMIN_INITIAL_BUDGET = 1_000_000

# ---- DB config (use environment variables in deployment) ----
DATABASE_URL = os.getenv("DATABASE_URL")  # Render provides this

if DATABASE_URL:
    # Parse DATABASE_URL: postgresql://user:pass@host:port/dbname
    import urllib.parse as urlparse
    urlparse.uses_netloc.append("postgresql")
    url = urlparse.urlparse(DATABASE_URL)
    DB_HOST = url.hostname
    DB_USER = url.username
    DB_PASS = url.password
    DB_NAME = url.path[1:]  # Remove leading /
    DB_PORT = url.port or 5432
else:
    # Local / fallback creds
    DB_HOST = os.getenv("PGHOST", "dpg-d426gaje5dus73bfka20-a.oregon-postgres.render.com")
    DB_USER = os.getenv("PGUSER", "admin_user")
    DB_PASS = os.getenv("PGPASSWORD", "NXUMDSA8WjBCkn5xBKFkxQGaKGaxNie8")
    DB_NAME = os.getenv("PGDATABASE", "grandguard")
    DB_PORT = int(os.getenv("PGPORT", "5432"))


def get_db():
    """Create a connection per request. No app-start crash if creds are wrong."""
    try:
        if DATABASE_URL:
            conn = psycopg2.connect(DATABASE_URL)
        else:
            conn = psycopg2.connect(
                host=DB_HOST,
                user=DB_USER,
                password=DB_PASS,
                database=DB_NAME,
                port=DB_PORT,
        )
        return conn
    except Exception as e:
        print(f"DB connect error: {e}")
        return None


def update_transactions_status_constraint():
    """Update transactions table constraint to include 'Paid' status."""
    conn = get_db()
    if conn is None:
        return False
    
    try:
        cur = conn.cursor()
        # Drop the old constraint if it exists
        cur.execute("""
            ALTER TABLE transactions
            DROP CONSTRAINT IF EXISTS transactions_status_check
        """)
        # Add the new constraint with 'Paid' status
        cur.execute("""
            ALTER TABLE transactions
            ADD CONSTRAINT transactions_status_check
            CHECK (status IN ('Pending', 'Approved', 'Paid', 'Declined'))
        """)
        conn.commit()
        cur.close()
        conn.close()
        return True
    except Exception as e:
        print(f"Error updating transactions status constraint: {e}")
        if conn:
            conn.rollback()
            conn.close()
        return False


def update_transactions_status_constraint():
    """Update transactions table constraint to include 'Paid' status."""
    conn = get_db()
    if conn is None:
        return False
    
    try:
        cur = conn.cursor()
        # Drop the old constraint if it exists
        cur.execute("""
            ALTER TABLE transactions
            DROP CONSTRAINT IF EXISTS transactions_status_check
        """)
        # Add the new constraint with 'Paid' status
        cur.execute("""
            ALTER TABLE transactions
            ADD CONSTRAINT transactions_status_check
            CHECK (status IN ('Pending', 'Approved', 'Paid', 'Declined'))
        """)
        conn.commit()
        cur.close()
        conn.close()
        print("✓ Transactions status constraint updated to include 'Paid'")
        return True
    except Exception as e:
        print(f"Error updating transactions status constraint: {e}")
        if conn:
            try:
                conn.rollback()
            except:
                pass
            conn.close()
        return False


def init_db_if_needed():
    """Initialize database schema if tables don't exist."""
    conn = get_db()
    if conn is None:
        print("Warning: Could not connect to database. Schema initialization skipped.")
        return

    try:
        cur = conn.cursor()
        schema_file = os.path.join(os.path.dirname(__file__), "schema_postgresql.sql")
        if os.path.exists(schema_file):
            with open(schema_file, 'r') as f:
                schema_sql = f.read()
            # Execute the entire schema file
            # PostgreSQL can handle multiple statements if we use execute with the full SQL
            try:
                # Remove comments and clean up
                lines = schema_sql.split('\n')
                cleaned_lines = []
                for line in lines:
                    # Skip comment lines and empty lines
                    stripped = line.strip()
                    if stripped and not stripped.startswith('--'):
                        cleaned_lines.append(line)
                cleaned_sql = '\n'.join(cleaned_lines)
                
                # Execute the cleaned SQL
                cur.execute(cleaned_sql)
                conn.commit()
                print("✓ Database schema initialized")
            except Exception as schema_error:
                # If full execution fails, try executing statement by statement
                print(f"Full schema execution failed, trying statement by statement: {schema_error}")
                statements = schema_sql.split(';')
                for statement in statements:
                    statement = statement.strip()
                    if statement and not statement.startswith('--') and not statement.upper().startswith('SELECT'):
                        try:
                            cur.execute(statement)
                        except Exception as stmt_error:
                            # Some statements might fail if tables already exist, that's okay
                            error_msg = str(stmt_error)
                            if 'already exists' not in error_msg.lower() and 'duplicate' not in error_msg.lower():
                                print(f"Statement execution note: {stmt_error}")
                conn.commit()
                print("✓ Database schema initialized (statement by statement)")
        else:
            print(f"Warning: Schema file {schema_file} not found")
        cur.close()
    except Exception as e:
        print(f"DB init error: {e}")
        import traceback
        traceback.print_exc()
        conn.rollback()
    finally:
        conn.close()
    
    # Update transactions status constraint after schema initialization
    update_transactions_status_constraint()


@app.route("/")
def home():
    return render_template("index.html")


# ========== Auth ==========

@app.route("/login", methods=["POST"])
def login():
    email = request.form.get("email", "").strip()
    password = request.form.get("password", "").strip()

    if not email or not password:
        return make_response("Missing credentials", 400)

    conn = get_db()
    if conn is None:
        return make_response("DB connection failed", 500)

    try:
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute(
            "SELECT name, role FROM users WHERE email=%s AND password=%s",
            (email, password),
        )
        user = cur.fetchone()
        cur.close()
        conn.close()
    except Exception as e:
        print(f"Query error: {e}")
        return make_response("DB query failed", 500)

    if not user:
        return make_response("Invalid email or password", 401)

    session["user"] = {"name": user["name"], "role": user["role"], "email": email}
    return "Welcome"


@app.route("/signup", methods=["POST"])
def signup():
    first = request.form.get("firstName", "").strip()
    last = request.form.get("lastName", "").strip()
    email = request.form.get("email", "").strip() or request.form.get("signupEmail", "").strip()
    password = request.form.get("password", "").strip() or request.form.get("signupPassword", "").strip()

    if not first or not last or not email or not password:
        return make_response("Missing required fields", 400)

    full_name = f"{first} {last}".strip()

    conn = get_db()
    if conn is None:
        return make_response("DB connection failed", 500)

    try:
        cur = conn.cursor()
        # Everyone who signs up is a PI by default
        cur.execute(
            """
            INSERT INTO users (name, email, role, password)
            VALUES (%s, %s, 'PI', %s)
            ON CONFLICT (email) DO UPDATE SET name = EXCLUDED.name
            """,
            (full_name, email, password),
        )
        conn.commit()
        cur.close()
    except Exception as e:
        print(f"DB insert user error: {e}")
        conn.rollback()
        return make_response(f"DB insert failed: {e}", 500)
    finally:
        conn.close()

    session["user"] = {"name": full_name, "role": "PI", "email": email}
    return "Signed up"


# ========== Dashboard (PI + Admin) ==========

@app.route("/dashboard")
def dashboard():
    """
    - If PI: show their own awards (PI dashboard).
    - If Admin: show all awards + budget info (Admin dashboard).
    """
    u = session.get("user")
    if not u:
        return render_template("dashboard.html", name="User", role="Guest", awards=[])

    # ---------- Admin dashboard ----------
    if u["role"] == "Admin":
        awards = []
        total_approved = 0.0
        conn = get_db()
        if conn is not None:
            try:
                cur = conn.cursor(cursor_factory=RealDictCursor)
                cur.execute(
                    """
                    SELECT award_id, title, created_by_email, sponsor_type,
                        amount, start_date, end_date, status, created_at, ai_review_notes
                    FROM awards
                    WHERE status <> 'Draft'
                    ORDER BY created_at DESC
                    """
                )
                awards = cur.fetchall()

                # Calculate total approved - sum all approved awards
                cur.execute(
                    """
                    SELECT COALESCE(SUM(amount), 0) as total
                    FROM awards 
                    WHERE status = 'Approved'
                    """
                )
                row = cur.fetchone()
                if row:
                    total_approved = float(row['total']) if row['total'] is not None else 0.0
                else:
                    total_approved = 0.0
                cur.close()
            except Exception as e:
                print(f"DB fetch awards (admin) error: {e}")
            finally:
                conn.close()

        budget_initial = float(ADMIN_INITIAL_BUDGET)
        budget_remaining = budget_initial - total_approved

        return render_template(
            "dashboard_admin.html",
            name=u["name"],
            role=u["role"],
            awards=awards,
            budget_initial=budget_initial,
            budget_remaining=budget_remaining,
        )

    # ---------- PI dashboard ----------
    awards = []
    conn = get_db()
    if conn is not None:
        try:
            cur = conn.cursor(cursor_factory=RealDictCursor)
            cur.execute(
                """
                SELECT award_id, title, sponsor_type, amount,
                       start_date, end_date, status, created_at
                FROM awards
                WHERE created_by_email=%s
                ORDER BY created_at DESC
                """,
                (u["email"],),
            )
            awards = cur.fetchall()
            cur.close()
        except Exception as e:
            print(f"DB fetch awards error: {e}")
        finally:
            conn.close()

    return render_template("dashboard.html", name=u["name"], role=u["role"], awards=awards)


# ========== Grants / Awards (PI side) ==========

@app.route("/awards/new")
def awards_new():
    """Show empty 'Generate Grants' form (PI only)."""
    u = session.get("user")
    if not u:
        return redirect(url_for("home"))
    if u["role"] != "PI":
        # Admins shouldn't create grants
        return redirect(url_for("dashboard"))
    return render_template("awards_new.html", award=None)

def _get_award_for_export(award_id, user):
    """
    Load a single award plus JSON budget fields and return:
    award, personnel, domestic_travel, international_travel, materials, equipment, other_direct
    Handles backward compatibility for equipment/other_direct stored in materials_json
    """
    conn = get_db()
    if conn is None:
        return None, [], [], [], [], [], []

    try:
        cur = conn.cursor(cursor_factory=RealDictCursor)

        # Admin can see all, PI only their own
        if user["role"] == "Admin":
            cur.execute("SELECT * FROM awards WHERE award_id=%s", (award_id,))
        else:
            cur.execute(
                "SELECT * FROM awards WHERE award_id=%s AND created_by_email=%s",
                (award_id, user["email"]),
            )

        award = cur.fetchone()
        cur.close()
        conn.close()
    except Exception as e:
        print(f"_get_award_for_export query error: {e}")
        return None, [], [], [], [], [], []

    if not award:
        return None, [], [], [], [], [], []

    # Parse JSON blobs exactly like award_view
    def parse_json(field_name):
        raw = award.get(field_name)
        if raw is None:
            return []
        if isinstance(raw, (dict, list)):
            return raw
        try:
            return json.loads(raw)
        except Exception:
            return []

    personnel = parse_json("personnel_json")
    domestic_travel = parse_json("domestic_travel_json")
    international_travel = parse_json("international_travel_json")
    materials_raw = parse_json("materials_json")
    equipment_from_json = parse_json("equipment_json")
    other_direct_from_json = parse_json("other_direct_json")
    
    # Extract equipment and other direct costs from materials_json (old format)
    # Equipment was stored in materials_json with type='equipment'
    equipment_from_materials = []
    materials = []
    other_direct_from_materials = []
    
    if materials_raw:
        for item in materials_raw:
            if isinstance(item, dict):
                item_type = item.get('type', '').lower()
                if item_type == 'equipment':
                    # Extract equipment from materials
                    equipment_from_materials.append({
                        'description': item.get('description', ''),
                        'cost': item.get('cost', 0)
                    })
                elif item_type == 'other':
                    # Extract other direct costs from materials
                    other_direct_from_materials.append({
                        'description': item.get('description', ''),
                        'cost': item.get('cost', 0)
                    })
                else:
                    # Regular materials
                    materials.append(item)
    
    # Combine equipment from both sources (equipment_json takes priority if it has items)
    if equipment_from_json and len(equipment_from_json) > 0:
        equipment = equipment_from_json
    else:
        equipment = equipment_from_materials
    
    # Combine other direct costs from both sources (other_direct_json takes priority if it has items)
    if other_direct_from_json and len(other_direct_from_json) > 0:
        other_direct = other_direct_from_json
    else:
        other_direct = other_direct_from_materials

    return award, personnel, domestic_travel, international_travel, materials, equipment, other_direct 

def _parse_json_field(field_value):
    """Helper: safely parse a JSON array field from the form."""
    if not field_value:
        return []
    try:
        data = json.loads(field_value)
        if isinstance(data, list):
            return data
        return []
    except json.JSONDecodeError:
        print("JSON decode error for field:", field_value[:200])
        return []


@app.route("/awards", methods=["POST"])
def awards_create():
    """Create a new award (PI submits; status defaults to 'Pending')."""
    u = session.get("user")
    if not u:
        return redirect(url_for("home"))
    if u["role"] != "PI":
        return redirect(url_for("dashboard"))

    title = request.form.get("title", "").strip()
    sponsor = None  # reserved
    sponsor_type = request.form.get("sponsor_type", "").strip()
    department = request.form.get("department", "").strip()
    college = request.form.get("college", "").strip()
    contact_email = request.form.get("contact_email", "").strip()
    amount = request.form.get("amount", "").strip()
    start_date = request.form.get("start_date", "").strip()
    end_date = request.form.get("end_date", "").strip()
    abstract = request.form.get("abstract", "").strip()
    keywords = request.form.get("keywords", "").strip()
    collaborators = request.form.get("collaborators", "").strip()

    # JSON strings for detailed budget sections
    personnel_json_str = request.form.get("personnel_json", "")
    domestic_travel_json_str = request.form.get("domestic_travel_json", "")
    international_travel_json_str = request.form.get("international_travel_json", "")
    materials_json_str = request.form.get("materials_json", "")
    equipment_json_str = request.form.get("equipment_json", "")
    other_costs_json_str = request.form.get("other_costs_json", "")  # Form uses other_costs_json, DB uses other_direct_json

    # Parse into Python lists (from the JS format)
    pers_list = _parse_json_field(personnel_json_str)
    dom_list = _parse_json_field(domestic_travel_json_str)
    intl_list = _parse_json_field(international_travel_json_str)
    mat_list = _parse_json_field(materials_json_str)
    equipment_list = _parse_json_field(equipment_json_str)
    other_direct_list = _parse_json_field(other_costs_json_str)

    if not title or not sponsor_type or not amount or not start_date or not end_date:
        return make_response("Missing required fields", 400)

    conn = get_db()
    if conn is None:
        return make_response("DB connection failed", 500)

    try:
        cur = conn.cursor()

        # Insert into awards and get award_id
        cur.execute(
            """
            INSERT INTO awards (
              created_by_email, title, sponsor, sponsor_type,
              department, college, contact_email,
              amount, start_date, end_date,
              abstract, keywords, collaborators,
              personnel_json, domestic_travel_json,
              international_travel_json, materials_json,
              equipment_json, other_direct_json
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s,
                    %s, %s, %s,
                    %s, %s, %s,
                    %s::jsonb, %s::jsonb, %s::jsonb, %s::jsonb,
                    %s::jsonb, %s::jsonb)
            RETURNING award_id
            """,
            (
                u["email"], title, sponsor, sponsor_type or None,
                department or None, college or None, contact_email or None,
                amount, start_date, end_date,
                abstract or None, keywords or None, collaborators or None,
                json.dumps(pers_list),
                json.dumps(dom_list),
                json.dumps(intl_list),
                json.dumps(mat_list),
                json.dumps(equipment_list),
                json.dumps(other_direct_list),
            ),
        )
        award_id = cur.fetchone()[0]

        # ---- Insert personnel rows (summary into personnel_expenses) ----
        for p in pers_list:
            name_val = (p.get("name") or "").strip()
            if not name_val:
                continue
            position = (p.get("position") or "").strip()
            same_each_year = bool(p.get("same_each_year", False))

            # p["hours"] is an array of {year, hours}
            hours_for_years = None
            hrs = p.get("hours")
            if isinstance(hrs, list) and hrs:
                total = 0.0
                for h in hrs:
                    try:
                        total += float(h.get("hours", 0) or 0)
                    except Exception:
                        pass
                if total > 0:
                    hours_for_years = total

            cur.execute(
                """
                INSERT INTO personnel_expenses
                  (award_id, person_name, position_title, hours_for_years, same_each_year)
                VALUES (%s, %s, %s, %s, %s)
                """,
                (award_id, name_val, position or None, hours_for_years, same_each_year),
            )

        # Helper for numeric conversion
        def num(v):
            try:
                return float(v) if v not in (None, "", "null") else None
            except (TypeError, ValueError):
                return None

        # ---- Insert domestic travel rows ----
        for t in dom_list:
            travel_name = (t.get("travel_name") or t.get("name") or "").strip()
            if not travel_name:
                continue
            desc = (t.get("description") or "").strip()
            year = t.get("year")
            start = t.get("start_date") or t.get("depart") or None
            end = t.get("end_date") or t.get("arrive") or None

            flight = num(t.get("flight_cost") or t.get("flight"))
            taxi = num(t.get("taxi_per_day"))
            food = num(t.get("food_lodge_per_day") or t.get("food_per_day"))
            days = None
            try:
                dval = t.get("days")
                days = int(dval) if dval not in (None, "", "null") else None
            except (TypeError, ValueError):
                days = None

            cur.execute(
                """
                INSERT INTO travel_expenses
                  (award_id, travel_type, travel_name, description, year,
                   start_date, end_date, flight_cost, taxi_per_day,
                   food_lodge_per_day, num_days)
                VALUES (%s, 'Domestic', %s, %s, %s,
                        %s, %s, %s, %s, %s, %s)
                """,
                (
                    award_id, travel_name, desc or None, year,
                    start, end, flight, taxi, food, days,
                ),
            )

        # ---- Insert international travel rows ----
        for t in intl_list:
            travel_name = (t.get("travel_name") or t.get("name") or "").strip()
            if not travel_name:
                continue
            desc = (t.get("description") or "").strip()
            year = t.get("year")
            start = t.get("start_date") or t.get("depart") or None
            end = t.get("end_date") or t.get("arrive") or None

            flight = num(t.get("flight_cost") or t.get("flight"))
            taxi = num(t.get("taxi_per_day"))
            food = num(t.get("food_lodge_per_day") or t.get("food_per_day"))
            days = None
            try:
                dval = t.get("days")
                days = int(dval) if dval not in (None, "", "null") else None
            except (TypeError, ValueError):
                days = None

            cur.execute(
                """
                INSERT INTO travel_expenses
                  (award_id, travel_type, travel_name, description, year,
                   start_date, end_date, flight_cost, taxi_per_day,
                   food_lodge_per_day, num_days)
                VALUES (%s, 'International', %s, %s, %s,
                        %s, %s, %s, %s, %s, %s)
                """,
                (
                    award_id, travel_name, desc or None, year,
                    start, end, flight, taxi, food, days,
                ),
            )

        # ---- Insert materials/supplies rows ----
        for m in mat_list:
            mtype = (m.get("material_type") or m.get("category") or "").strip()
            if not mtype:
                continue
            desc = (m.get("description") or "").strip()
            year = m.get("year")
            try:
                cost_val = float(m.get("cost")) if m.get("cost") not in (None, "", "null") else None
            except (TypeError, ValueError):
                cost_val = None

            cur.execute(
                """
                INSERT INTO material_supplies
                  (award_id, material_type, cost, description, year)
                VALUES (%s, %s, %s, %s, %s)
                """,
                (award_id, mtype, cost_val, desc or None, year),
            )

        conn.commit()
        cur.close()
    except Exception as e:
        print(f"DB insert award error: {e}")
        conn.rollback()
        return make_response(f"DB insert failed: {e}", 500)
    finally:
        conn.close()

    return redirect(url_for("dashboard"))


@app.route("/awards/<int:award_id>/view")
def award_view(award_id):
    u = session.get("user")
    if not u:
        return redirect(url_for("home"))

    conn = get_db()
    if conn is None:
        return make_response("DB connection failed", 500)

    try:
        cur = conn.cursor(cursor_factory=RealDictCursor)

        # Admin can see all, PI only their own
        # Explicitly select all columns including JSONB fields
        if u["role"] == "Admin":
            cur.execute("""
                SELECT *, 
                       equipment_json::text as equipment_json_text,
                       other_direct_json::text as other_direct_json_text
                FROM awards WHERE award_id=%s
            """, (award_id,))
        else:
            cur.execute(
                """
                SELECT *, 
                       equipment_json::text as equipment_json_text,
                       other_direct_json::text as other_direct_json_text
                FROM awards WHERE award_id=%s AND created_by_email=%s
                """,
                (award_id, u["email"]),
            )

        award = cur.fetchone()
        cur.close()
        conn.close()
    except Exception as e:
        print(f"DB fetch single award error: {e}")
        return make_response("DB query failed", 500)

    if not award:
        return "Award not found", 404

    # --- Parse JSON blobs safely ---
    def parse_json(field_name):
        # Try both the original field and the text version
        raw = award.get(field_name)
        if raw is None:
            # Try the text version if available
            text_field = field_name + "_text"
            raw = award.get(text_field)
            if raw is None:
                return []
        
        # RealDictCursor might return JSONB as dict/list already, or as string
        if isinstance(raw, (dict, list)):
            # If it's already a dict/list, return it (but ensure it's a list)
            if isinstance(raw, list):
                return raw
            # If it's a dict, wrap it in a list (shouldn't happen for arrays but handle it)
            return [raw] if raw else []
        
        # If it's a string, try to parse it
        if isinstance(raw, str):
            try:
                parsed = json.loads(raw)
                # Ensure we return a list
                if isinstance(parsed, list):
                    return parsed
                elif isinstance(parsed, dict):
                    # If it's a dict, wrap it in a list
                    return [parsed]
                return []
            except (json.JSONDecodeError, TypeError) as e:
                print(f"Error parsing {field_name}: {e}, raw value: {raw}")
                return []
        
        return []

    personnel = parse_json("personnel_json")
    # Ensure rate_per_hour and total are floats for display
    for p in personnel:
        if isinstance(p, dict):
            if 'rate_per_hour' in p:
                try:
                    p['rate_per_hour'] = float(p['rate_per_hour'] or 0)
                except (ValueError, TypeError):
                    p['rate_per_hour'] = 0.0
            if 'total' in p:
                try:
                    p['total'] = float(p['total'] or 0)
                except (ValueError, TypeError):
                    p['total'] = 0.0
    
    domestic_travel = parse_json("domestic_travel_json")
    international_travel = parse_json("international_travel_json")
    materials_raw = parse_json("materials_json")
    equipment_from_json = parse_json("equipment_json")
    other_direct_from_json = parse_json("other_direct_json")
    
    # Extract equipment and other direct costs from materials_json (old format)
    # Equipment was stored in materials_json with type='equipment'
    equipment_from_materials = []
    materials = []
    other_direct_from_materials = []
    
    if materials_raw:
        for item in materials_raw:
            if isinstance(item, dict):
                item_type = item.get('type', '').lower()
                if item_type == 'equipment':
                    # Extract equipment from materials
                    equipment_from_materials.append({
                        'description': item.get('description', ''),
                        'cost': item.get('cost', 0)
                    })
                elif item_type == 'other':
                    # Extract other direct costs from materials
                    other_direct_from_materials.append({
                        'description': item.get('description', ''),
                        'cost': item.get('cost', 0)
                    })
                else:
                    # Regular materials
                    materials.append(item)
    
    # Combine equipment from both sources (equipment_json takes priority if it has items)
    if equipment_from_json and len(equipment_from_json) > 0:
        equipment = equipment_from_json
    else:
        equipment = equipment_from_materials
    
    # Combine other direct costs from both sources (other_direct_json takes priority if it has items)
    if other_direct_from_json and len(other_direct_from_json) > 0:
        other_direct = other_direct_from_json
    else:
        other_direct = other_direct_from_materials

    # --- Compute period & year list for tables ---
    start = award.get("start_date")
    end = award.get("end_date")

    years = []
    duration_years = None
    if isinstance(start, date) and isinstance(end, date) and end >= start:
        # Calculate actual duration in years more accurately
        # If end date is Jan 1 of the next year after start, it's exactly 1 year
        if end.year == start.year + 1 and end.month == 1 and end.day == 1 and start.month == 1 and start.day == 1:
            # Exactly one year (e.g., 2026-01-01 to 2027-01-01)
            duration_years = 1
            years = [start.year]
        else:
            # Calculate years more accurately
            years = list(range(start.year, end.year + 1))
            # Calculate duration: if end date is at the start of a year, don't count that full year
            if end.month == 1 and end.day == 1:
                # End date is Jan 1, so the period doesn't include that year
                duration_years = end.year - start.year
                years = list(range(start.year, end.year))
            else:
                # Include both start and end years
                duration_years = end.year - start.year + 1

    # Load AI compliance results if they exist
    compliance_results = None
    if award.get('ai_review_notes'):
        try:
            compliance_results = json.loads(award['ai_review_notes'])
        except (json.JSONDecodeError, TypeError):
            compliance_results = None

    return render_template(
        "award_view.html",
        award=award,
        personnel=personnel,
        domestic_travel=domestic_travel,
        international_travel=international_travel,
        materials=materials,
        equipment=equipment,
        other_direct=other_direct,
        years=years,
        duration_years=duration_years,
        compliance_results=compliance_results,
        user=u,
    )


# ========== EXPORTS: Excel + PDF ==========

@app.route("/awards/<int:award_id>/download/pdf")
def download_award_pdf(award_id):
    u = session.get("user")
    if not u:
        return redirect(url_for("home"))

    award, personnel, domestic_travel, international_travel, materials, equipment, other_direct = _get_award_for_export(award_id, u)
    if not award:
        return "Award not found", 404

    # -------- helpers ----------
    def hours_text(hours_list):
        if not hours_list:
            return ""
        parts = []
        for h in hours_list:
            year = h.get("year")
            hrs = h.get("hours")
            if year and hrs not in (None, ""):
                parts.append(f"{year}: {hrs} hrs")
        return ", ".join(parts)

    def travel_row(travel_type, t):
        # Handle new format (total_amount) and old format (flight/taxi/food)
        total_amount = float(t.get("total_amount") or 0)
        if total_amount > 0:
            return [
                travel_type,
                t.get("description") or "",
                f"${total_amount:,.2f}",
            ]
        else:
            # Old format fallback
            flight = float(t.get("flight_cost") or t.get("flight") or 0)
            taxi = float(t.get("taxi_per_day") or 0)
            food = float(t.get("food_lodge_per_day") or t.get("food_per_day") or 0)
            days = float(t.get("days") or 0)
            old_total = flight + (taxi + food) * days if days > 0 else flight
            return [
                travel_type,
                t.get("description") or "",
                f"${old_total:,.2f}" if old_total > 0 else "",
            ]

    # -------- basic fields ----------
    title = award.get("title") or "Grant"
    funding = award.get("sponsor_type") or "N/A"
    amount = float(award.get("amount") or 0)
    dept = award.get("department") or "N/A"
    college = award.get("college") or "N/A"
    email = award.get("contact_email") or award.get("created_by_email") or "N/A"
    status = award.get("status") or "Pending"
    start = award.get("start_date")
    end = award.get("end_date")
    abstract = award.get("abstract") or "N/A"
    keywords = award.get("keywords") or "N/A"
    collaborators = award.get("collaborators") or "N/A"

    if start and end:
        period_str = f"{start} \u2192 {end}"
    else:
        period_str = "N/A"

    # -------- build the PDF with tables ----------
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        leftMargin=36,
        rightMargin=36,
        topMargin=36,
        bottomMargin=36,
    )
    styles = getSampleStyleSheet()
    normal = styles["Normal"]
    title_style = styles["Title"]

    elements = []

    # Top title
    elements.append(Paragraph(title, title_style))
    elements.append(Spacer(1, 8))

    # Summary block – similar to the top of the HTML view
    summary_lines = [
        f"<b>Funding Agency:</b> {funding}",
        f"<b>Amount:</b> ${amount:,.2f}",
        f"<b>Period:</b> {period_str}",
        f"<b>Status:</b> {status}",
        f"<b>Department:</b> {dept}",
        f"<b>College:</b> {college}",
        f"<b>Contact Email:</b> {email}",
    ]
    for line in summary_lines:
        elements.append(Paragraph(line, normal))
    elements.append(Spacer(1, 10))

    # Abstract / keywords / collaborators
    elements.append(Paragraph("<b>Abstract:</b>", normal))
    elements.append(Paragraph(abstract, normal))
    elements.append(Spacer(1, 6))
    elements.append(Paragraph(f"<b>Keywords:</b> {keywords}", normal))
    elements.append(Paragraph(f"<b>Collaborators:</b> {collaborators}", normal))
    elements.append(Spacer(1, 12))

    # -------- Personnel table ----------
    if personnel:
        elements.append(Paragraph("Personnel", styles["Heading3"]))
        data = [["Name", "Position", "Hours for year(s)", "Rate per Hour", "Total Amount", "Same Each Year?"]]
        for p in personnel:
            rate = float(p.get("rate_per_hour") or 0)
            total = float(p.get("total") or 0)
            # Calculate total if not provided
            if total == 0 and rate > 0:
                hours_list = p.get("hours", [])
                if isinstance(hours_list, list):
                    total_hours = sum(float(h.get("hours", 0) or 0) for h in hours_list if isinstance(h, dict))
                    total = total_hours * rate
            
            data.append([
                p.get("name") or "",
                p.get("position") or "",
                hours_text(p.get("hours")),
                f"${rate:,.2f}" if rate > 0 else "",
                f"${total:,.2f}" if total > 0 else "",
                "Yes" if p.get("same_each_year") else "No",
            ])

        t = Table(data, repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E0E0E0")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ("ALIGN", (3, 1), (4, -1), "RIGHT"),  # Right-align rate and total columns
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))
        elements.append(t)
        elements.append(Spacer(1, 12))

    # -------- Travel table (domestic + international) ----------
    if domestic_travel or international_travel:
        elements.append(Paragraph("Travel Information", styles["Heading3"]))
        data = [
            [
                "Type", "Description", "Total Estimated Amount"
            ]
        ]
        for t_dom in domestic_travel:
            data.append(travel_row("Domestic", t_dom))
        for t_int in international_travel:
            data.append(travel_row("International", t_int))

        t = Table(data, repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E0E0E0")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ("ALIGN", (2, 1), (2, -1), "RIGHT"),  # Right-align amount column
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))
        elements.append(t)
        elements.append(Spacer(1, 12))

    # -------- Materials table ----------
    if materials:
        elements.append(Paragraph("Materials and Supplies", styles["Heading3"]))
        data = [["Description", "Cost"]]
        for m in materials:
            cost = float(m.get("cost") or 0)
            data.append([
                m.get("description") or "",
                f"${cost:,.2f}" if cost > 0 else "",
            ])

        t = Table(data, repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E0E0E0")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ("ALIGN", (1, 1), (1, -1), "RIGHT"),  # Right-align cost column
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))
        elements.append(t)
        elements.append(Spacer(1, 12))
    
    # -------- Equipment ----------
    if equipment:
        elements.append(Paragraph("Equipment", styles["Heading3"]))
        data = [["Description", "Cost"]]

        for e in equipment:
            cost = float(e.get("cost") or 0)
            data.append([
                e.get("description") or "",
                f"${cost:,.2f}" if cost > 0 else "",
            ])

        t = Table(data, repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E0E0E0")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ("ALIGN", (1, 1), (1, -1), "RIGHT"),  # Right-align cost column
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))
        elements.append(t)
        elements.append(Spacer(1, 12))
    
    # -------- Other Direct Costs ----------
    if other_direct:
        elements.append(Paragraph("Other Direct Costs", styles["Heading3"]))
        data = [["Description", "Cost"]]

        for d in other_direct:
            cost = float(d.get("cost") or 0)
            data.append([
                d.get("description") or "",
                f"${cost:,.2f}" if cost > 0 else "",
            ])

        t = Table(data, repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E0E0E0")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ("ALIGN", (1, 1), (1, -1), "RIGHT"),  # Right-align cost column
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))
        elements.append(t)
    # Build PDF
    doc.build(elements)
    buffer.seek(0)

    filename = f"grant_{award.get('award_id', award_id)}.pdf"
    resp = make_response(buffer.getvalue())
    resp.headers["Content-Type"] = "application/pdf"
    resp.headers["Content-Disposition"] = f'attachment; filename=\"{filename}\"'
    return resp

@app.route("/awards/<int:award_id>/download/excel")
def download_award_excel(award_id):
    u = session.get("user")
    if not u:
        return redirect(url_for("home"))

    award, personnel, domestic_travel, international_travel, materials, equipment, other_direct = _get_award_for_export(award_id, u)
    if not award:
        return "Award not found", 404

    wb = Workbook()
    ws = wb.active
    ws.title = "Grant Budget"

    # Styles
    header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    bold = Font(bold=True)
    center = Alignment(horizontal="center")
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Set some decent column widths
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 12
    ws.column_dimensions["I"].width = 16
    ws.column_dimensions["J"].width = 10

    row = 1

    def write_kv(label, value):
        nonlocal row
        ws.cell(row=row, column=1, value=label).font = bold
        ws.cell(row=row, column=2, value=value)
        row += 1

    title = award.get("title") or ""
    funding = award.get("sponsor_type") or ""
    amount = float(award.get("amount") or 0)
    dept = award.get("department") or ""
    college = award.get("college") or ""
    email = award.get("contact_email") or award.get("created_by_email") or ""
    status = award.get("status") or ""
    start = award.get("start_date")
    end = award.get("end_date")
    abstract = award.get("abstract") or ""
    keywords = award.get("keywords") or ""
    collaborators = award.get("collaborators") or ""

    write_kv("Title", title)
    write_kv("Funding Agency", funding)
    write_kv("Amount", amount)
    write_kv("Department", dept)
    write_kv("College", college)
    write_kv("Contact Email", email)
    write_kv("Start Date", start)
    write_kv("End Date", end)
    write_kv("Status", status)
    row += 1
    write_kv("Abstract", abstract)
    write_kv("Keywords", keywords)
    write_kv("Collaborators", collaborators)

    row += 2

    # Personnel section
    if personnel:
        ws.cell(row=row, column=1, value="Personnel").font = bold
        row += 1
        headers = ["Name", "Position", "Hours by Year", "Rate per Hour", "Total Amount", "Same Each Year?"]
        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = bold
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border
        row += 1

        def hours_text(hours_list):
            if not hours_list:
                return ""
            parts = []
            for h in hours_list:
                year = h.get("year")
                hrs = h.get("hours")
                if year and hrs not in (None, ""):
                    parts.append(f"{year}: {hrs} hrs")
            return ", ".join(parts)

        for p in personnel:
            rate = float(p.get("rate_per_hour") or 0)
            total = float(p.get("total") or 0)
            # Calculate total if not provided
            if total == 0 and rate > 0:
                hours_list = p.get("hours", [])
                if isinstance(hours_list, list):
                    total_hours = sum(float(h.get("hours", 0) or 0) for h in hours_list if isinstance(h, dict))
                    total = total_hours * rate
            
            ws.cell(row=row, column=1, value=p.get("name") or "").border = border
            ws.cell(row=row, column=2, value=p.get("position") or "").border = border
            ws.cell(row=row, column=3, value=hours_text(p.get("hours"))).border = border
            ws.cell(row=row, column=4, value=rate if rate > 0 else "").border = border
            if rate > 0:
                ws.cell(row=row, column=4).number_format = '$#,##0.00'
            ws.cell(row=row, column=5, value=total if total > 0 else "").border = border
            if total > 0:
                ws.cell(row=row, column=5).number_format = '$#,##0.00'
            ws.cell(
                row=row,
                column=6,
                value="Yes" if p.get("same_each_year") else "No",
            ).border = border
            row += 1
        row += 2

    # Travel section (domestic + international)
    if domestic_travel or international_travel:
        ws.cell(row=row, column=1, value="Travel").font = bold
        row += 1
        headers = [
            "Type", "Description", "Total Estimated Amount"
        ]
        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = bold
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border
        row += 1

        def add_travel_row(travel_type, t):
            nonlocal row
            # Handle new format (total_amount) and old format (flight/taxi/food)
            total_amount = float(t.get("total_amount") or 0)
            if total_amount > 0:
                ws.cell(row=row, column=1, value=travel_type).border = border
                ws.cell(row=row, column=2, value=t.get("description") or "").border = border
                ws.cell(row=row, column=3, value=total_amount).border = border
                ws.cell(row=row, column=3).number_format = '$#,##0.00'
            else:
                # Old format fallback
                flight = float(t.get("flight_cost") or t.get("flight") or 0)
                taxi = float(t.get("taxi_per_day") or 0)
                food = float(t.get("food_lodge_per_day") or t.get("food_per_day") or 0)
                days = float(t.get("days") or 0)
                old_total = flight + (taxi + food) * days if days > 0 else flight
                ws.cell(row=row, column=1, value=travel_type).border = border
                ws.cell(row=row, column=2, value=t.get("description") or "").border = border
                if old_total > 0:
                    ws.cell(row=row, column=3, value=old_total).border = border
                    ws.cell(row=row, column=3).number_format = '$#,##0.00'
                else:
                    ws.cell(row=row, column=3, value="").border = border
            row += 1

        for t in domestic_travel:
            add_travel_row("Domestic", t)
        for t in international_travel:
            add_travel_row("International", t)
        row += 2

    # Materials section
    if materials:
        ws.cell(row=row, column=1, value="Materials and Supplies").font = bold
        row += 1
        headers = ["Description", "Cost"]
        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = bold
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border
        row += 1

        for m in materials:
            cost = float(m.get("cost") or 0)
            ws.cell(row=row, column=1, value=m.get("description") or "").border = border
            if cost > 0:
                ws.cell(row=row, column=2, value=cost).border = border
                ws.cell(row=row, column=2).number_format = '$#,##0.00'
            else:
                ws.cell(row=row, column=2, value="").border = border
            row += 1
        row += 1
    
    # Equipment section
    if equipment:
        ws.cell(row=row, column=1, value="Equipment").font = bold
        row += 1
        headers = ["Description", "Cost"]

        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = bold
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border
        row += 1

        for e in equipment:
            cost = float(e.get("cost") or 0)
            ws.cell(row=row, column=1, value=e.get("description") or "").border = border
            if cost > 0:
                ws.cell(row=row, column=2, value=cost).border = border
                ws.cell(row=row, column=2).number_format = '$#,##0.00'
            else:
                ws.cell(row=row, column=2, value="").border = border
            row += 1
        row += 1
    
    # Other Direct Costs
    if other_direct:
        ws.cell(row=row, column=1, value="Other Direct Costs").font = bold
        row += 1
        headers = ["Description", "Cost"]

        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = bold
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border
        row += 1

        for d in other_direct:
            cost = float(d.get("cost") or 0)
            ws.cell(row=row, column=1, value=d.get("description") or "").border = border
            if cost > 0:
                ws.cell(row=row, column=2, value=cost).border = border
                ws.cell(row=row, column=2).number_format = '$#,##0.00'
            else:
                ws.cell(row=row, column=2, value="").border = border
            row += 1
    # Save to memory and return
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    filename = f"grant_{award.get('award_id', award_id)}.xlsx"
    resp = make_response(bio.getvalue())
    resp.headers["Content-Type"] = (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp.headers["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp

# ========== Edit / Delete / Submit / Admin ==========

@app.route("/awards/<int:award_id>/edit", methods=["GET", "POST"])
def award_edit(award_id):
    """Edit an existing award (PI only). Reuses awards_new.html."""
    u = session.get("user")
    if not u:
        return redirect(url_for("home"))
    if u["role"] != "PI":
        return redirect(url_for("dashboard"))

    conn = get_db()
    if conn is None:
        return make_response("DB connection failed", 500)

    if request.method == "POST":
        title = request.form.get("title", "").strip()
        sponsor_type = request.form.get("sponsor_type", "").strip()
        department = request.form.get("department", "").strip()
        college = request.form.get("college", "").strip()
        contact_email = request.form.get("contact_email", "").strip()
        amount = request.form.get("amount", "").strip()
        start_date = request.form.get("start_date", "").strip()
        end_date = request.form.get("end_date", "").strip()
        abstract = request.form.get("abstract", "").strip()
        keywords = request.form.get("keywords", "").strip()
        collaborators = request.form.get("collaborators", "").strip()

        personnel_json_str = request.form.get("personnel_json", "")
        domestic_travel_json_str = request.form.get("domestic_travel_json", "")
        international_travel_json_str = request.form.get("international_travel_json", "")
        materials_json_str = request.form.get("materials_json", "")
        equipment_json_str = request.form.get("equipment_json", "")
        other_costs_json_str = request.form.get("other_costs_json", "")  # Form uses other_costs_json, DB uses other_direct_json

        pers_list = _parse_json_field(personnel_json_str)
        dom_list = _parse_json_field(domestic_travel_json_str)
        intl_list = _parse_json_field(international_travel_json_str)
        mat_list = _parse_json_field(materials_json_str)
        equipment_list = _parse_json_field(equipment_json_str)
        other_direct_list = _parse_json_field(other_costs_json_str)

        if not title or not sponsor_type or not amount or not start_date or not end_date:
            return make_response("Missing required fields", 400)

        try:
            cur = conn.cursor()
            
            # Ensure equipment_json and other_direct_json columns exist (if they don't, this will fail silently)
            try:
                cur.execute("ALTER TABLE awards ADD COLUMN IF NOT EXISTS equipment_json JSONB")
            except Exception:
                pass  # Column might already exist
            try:
                cur.execute("ALTER TABLE awards ADD COLUMN IF NOT EXISTS other_direct_json JSONB")
            except Exception:
                pass  # Column might already exist

            # Update master award + JSON blobs
            cur.execute(
                """
                UPDATE awards
                SET title=%s,
                    sponsor_type=%s,
                    department=%s,
                    college=%s,
                    contact_email=%s,
                    amount=%s,
                    start_date=%s,
                    end_date=%s,
                    abstract=%s,
                    keywords=%s,
                    collaborators=%s,
                    personnel_json=%s::jsonb,
                    domestic_travel_json=%s::jsonb,
                    international_travel_json=%s::jsonb,
                    materials_json=%s::jsonb,
                    equipment_json=%s::jsonb,
                    other_direct_json=%s::jsonb
                WHERE award_id=%s AND created_by_email=%s
                """,
                (
                    title, sponsor_type or None,
                    department or None, college or None, contact_email or None,
                    amount, start_date, end_date,
                    abstract or None, keywords or None, collaborators or None,
                    json.dumps(pers_list),
                    json.dumps(dom_list),
                    json.dumps(intl_list),
                    json.dumps(mat_list),
                    json.dumps(equipment_list),
                    json.dumps(other_direct_list),
                    award_id, u["email"],
                ),
            )

            # Wipe existing detail rows and re-insert
            cur.execute("DELETE FROM personnel_expenses WHERE award_id=%s", (award_id,))
            cur.execute("DELETE FROM travel_expenses WHERE award_id=%s", (award_id,))
            cur.execute("DELETE FROM material_supplies WHERE award_id=%s", (award_id,))

            # Helper for numeric
            def num(v):
                try:
                    return float(v) if v not in (None, "", "null") else None
                except (TypeError, ValueError):
                    return None

            # Re-insert personnel
            for p in pers_list:
                name_val = (p.get("name") or "").strip()
                if not name_val:
                    continue
                position = (p.get("position") or "").strip()
                same_each_year = bool(p.get("same_each_year", False))

                hours_for_years = None
                hrs = p.get("hours")
                if isinstance(hrs, list) and hrs:
                    total = 0.0
                    for h in hrs:
                        try:
                            total += float(h.get("hours", 0) or 0)
                        except Exception:
                            pass
                    if total > 0:
                        hours_for_years = total

                cur.execute(
                    """
                    INSERT INTO personnel_expenses
                      (award_id, person_name, position_title, hours_for_years, same_each_year)
                    VALUES (%s, %s, %s, %s, %s)
                    """,
                    (award_id, name_val, position or None, hours_for_years, same_each_year),
                )

            # Re-insert domestic travel
            for t in dom_list:
                travel_name = (t.get("travel_name") or t.get("name") or "").strip()
                if not travel_name:
                    continue
                desc = (t.get("description") or "").strip()
                year = t.get("year")
                start = t.get("start_date") or t.get("depart") or None
                end = t.get("end_date") or t.get("arrive") or None

                flight = num(t.get("flight_cost") or t.get("flight"))
                taxi = num(t.get("taxi_per_day"))
                food = num(t.get("food_lodge_per_day") or t.get("food_per_day"))
                days = None
                try:
                    dval = t.get("days")
                    days = int(dval) if dval not in (None, "", "null") else None
                except (TypeError, ValueError):
                    days = None

                cur.execute(
                    """
                    INSERT INTO travel_expenses
                      (award_id, travel_type, travel_name, description, year,
                       start_date, end_date, flight_cost, taxi_per_day,
                       food_lodge_per_day, num_days)
                    VALUES (%s, 'Domestic', %s, %s, %s,
                            %s, %s, %s, %s, %s, %s)
                    """,
                    (
                        award_id, travel_name, desc or None, year,
                        start, end, flight, taxi, food, days,
                    ),
                )

            # Re-insert international travel
            for t in intl_list:
                travel_name = (t.get("travel_name") or t.get("name") or "").strip()
                if not travel_name:
                    continue
                desc = (t.get("description") or "").strip()
                year = t.get("year")
                start = t.get("start_date") or t.get("depart") or None
                end = t.get("end_date") or t.get("arrive") or None

                flight = num(t.get("flight_cost") or t.get("flight"))
                taxi = num(t.get("taxi_per_day"))
                food = num(t.get("food_lodge_per_day") or t.get("food_per_day"))
                days = None
                try:
                    dval = t.get("days")
                    days = int(dval) if dval not in (None, "", "null") else None
                except (TypeError, ValueError):
                    days = None

                cur.execute(
                    """
                    INSERT INTO travel_expenses
                      (award_id, travel_type, travel_name, description, year,
                       start_date, end_date, flight_cost, taxi_per_day,
                       food_lodge_per_day, num_days)
                    VALUES (%s, 'International', %s, %s, %s,
                            %s, %s, %s, %s, %s, %s)
                    """,
                    (
                        award_id, travel_name, desc or None, year,
                        start, end, flight, taxi, food, days,
                    ),
                )

            # Re-insert materials/supplies
            for m in mat_list:
                mtype = (m.get("material_type") or m.get("category") or "").strip()
                if not mtype:
                    continue
                desc = (m.get("description") or "").strip()
                year = m.get("year")
                try:
                    cost_val = float(m.get("cost")) if m.get("cost") not in (None, "", "null") else None
                except (TypeError, ValueError):
                    cost_val = None

                cur.execute(
                    """
                    INSERT INTO material_supplies
                      (award_id, material_type, cost, description, year)
                    VALUES (%s, %s, %s, %s, %s)
                    """,
                    (award_id, mtype, cost_val, desc or None, year),
                )

            conn.commit()
            cur.close()
            
            # Recalculate budget lines if award is approved (to reflect updated form data)
            cur = conn.cursor(cursor_factory=RealDictCursor)
            cur.execute("SELECT status FROM awards WHERE award_id = %s", (award_id,))
            award_status = cur.fetchone()
            cur.close()
            
            if award_status and award_status['status'] == 'Approved':
                initialize_budget_lines(award_id)
                
        except Exception as e:
            print(f"DB update award error: {e}")
            conn.rollback()
            return make_response("Update failed", 500)
        finally:
            conn.close()

        return redirect(url_for("dashboard"))

    # GET: load existing award
    try:
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute(
            "SELECT * FROM awards WHERE award_id=%s AND created_by_email=%s",
            (award_id, u["email"]),
        )
        award = cur.fetchone()
        cur.close()
        conn.close()
    except Exception as e:
        print(f"DB fetch single award error: {e}")
        return make_response("DB query failed", 500)

    if not award:
        return "Award not found", 404

    # For now we only prefill the master fields (form JS could later prefill JSON details)
    return render_template("awards_new.html", award=award)


@app.route("/awards/<int:award_id>/delete", methods=["POST"])
def award_delete(award_id):
    """Delete an award.

    - PIs can delete their own awards.
    - Admins can delete awards that are already declined.
    """
    u = session.get("user")
    if not u:
        return redirect(url_for("home"))

    conn = get_db()
    if conn is None:
        return make_response("DB connection failed", 500)

    try:
        cur = conn.cursor(cursor_factory=RealDictCursor)

        # Fetch the award with role-aware constraints
        if u["role"] == "PI":
            cur.execute(
                "SELECT award_id, status FROM awards WHERE award_id=%s AND created_by_email=%s",
                (award_id, u["email"]),
            )
        else:
            # Admin: can delete only if the award is declined
            cur.execute(
                "SELECT award_id, status FROM awards WHERE award_id=%s",
                (award_id,),
            )

        award = cur.fetchone()

        if not award:
            cur.close()
            conn.close()
            return "Award not found", 404

        # Authorization rules
        if u["role"] == "PI":
            # PIs can delete their own awards (any status for now)
            pass
        elif u["role"] == "Admin":
            if (award.get("status") or "").lower() != "declined":
                cur.close()
                conn.close()
                return make_response("Admins may only delete declined awards.", 403)
        else:
                cur.close()
                conn.close()
                # Update transactions status constraint
                update_transactions_status_constraint()
                return redirect(url_for("dashboard"))

        # Perform delete
        cur.execute(
            "DELETE FROM awards WHERE award_id=%s",
            (award_id,),
        )
        conn.commit()
        cur.close()
    except Exception as e:
        print(f"DB delete award error: {e}")
        conn.rollback()
        return make_response("Delete failed", 500)
    finally:
        conn.close()

    return redirect(url_for("dashboard"))


@app.route("/awards/<int:award_id>/submit", methods=["POST"])
def award_submit(award_id):
    """
    PI clicks Submit on dashboard – mark award as 'Pending'.
    """
    u = session.get("user")
    if not u:
        return redirect(url_for("home"))
    if u["role"] != "PI":
        return redirect(url_for("dashboard"))

    conn = get_db()
    if conn is None:
        return make_response("DB connection failed", 500)

    try:
        cur = conn.cursor()
        cur.execute(
            "UPDATE awards SET status = %s WHERE award_id = %s AND created_by_email = %s",
            ("Pending", award_id, u["email"]),
        )
        conn.commit()
        cur.close()
    except Exception as e:
        print(f"DB submit award error: {e}")
        conn.rollback()
        return make_response("Submit failed", 500)
    finally:
        conn.close()

    return redirect(url_for("dashboard"))


# ========== Admin actions: approve / decline ==========

@app.route("/awards/<int:award_id>/approve", methods=["POST"])
def award_approve(award_id):
    u = session.get("user")
    if not u or u.get("role") != "Admin":
        return redirect(url_for("home"))

    conn = get_db()
    if conn is None:
        return make_response("DB connection failed", 500)

    try:
        cur = conn.cursor(cursor_factory=RealDictCursor)

        # Amount of this award
        cur.execute("SELECT amount FROM awards WHERE award_id=%s", (award_id,))
        row = cur.fetchone()
        if not row:
            cur.close()
            conn.close()
            return "Award not found", 404

        amount_val = row['amount']
        amount = float(amount_val) if amount_val is not None else 0.0

        # Current approved total (EXCLUDE the current award being approved)
        cur.execute(
            """
            SELECT COALESCE(SUM(amount), 0) as total
            FROM awards 
            WHERE status = 'Approved' AND award_id != %s
            """,
            (award_id,)
        )
        row = cur.fetchone()
        if row:
            total_approved = float(row['total']) if row['total'] is not None else 0.0
        else:
            total_approved = 0.0
        remaining = float(ADMIN_INITIAL_BUDGET) - total_approved

        if remaining < amount:
            cur.close()
            conn.close()
            return make_response(f"Not enough remaining admin budget to approve this award. Remaining: ${remaining:,.2f}, Required: ${amount:,.2f}", 400)

        # Update status to Approved
        cur.execute(
            "UPDATE awards SET status='Approved' WHERE award_id=%s",
            (award_id,),
        )
        conn.commit()
                        # Initialize budget lines for approved award (must happen after commit)
        initialize_budget_lines(award_id)
        cur.close()
        conn.close()
    except Exception as e:
        print(f"DB approve award error: {e}")
        conn.rollback()
        return make_response("Approve failed", 500)
    finally:
        conn.close()

    return redirect(url_for("dashboard"))


@app.route("/awards/<int:award_id>/decline", methods=["POST"])
def award_decline(award_id):
    u = session.get("user")
    if not u or u.get("role") != "Admin":
        return redirect(url_for("home"))

    conn = get_db()
    if conn is None:
        return make_response("DB connection failed", 500)

    try:
        cur = conn.cursor()
        cur.execute(
            "UPDATE awards SET status='Declined' WHERE award_id=%s",
            (award_id,),
        )
        conn.commit()
        cur.close()
    except Exception as e:
        print(f"DB decline award error: {e}")
        conn.rollback()
        return make_response("Decline failed", 500)
    finally:
        conn.close()

    return redirect(url_for("dashboard"))


# ========== Other pages ==========

# ========== SUBAWARDS SYSTEM ==========

@app.route("/subawards")
def subawards():
    """List all subawards for the user."""
    u = session.get("user")
    if not u:
        return redirect(url_for("home"))
    
    conn = get_db()
    subawards_list = []
    awards_map = {}
    
    if conn is not None:
        try:
            cur = conn.cursor(cursor_factory=RealDictCursor)
            
            if u["role"] == "Admin":
                # Admin sees all subawards
                cur.execute(
                    """
                    SELECT s.*, a.title as award_title, a.status as award_status
                    FROM subawards s
                    LEFT JOIN awards a ON s.award_id = a.award_id
                    ORDER BY s.created_at DESC
                    """
                )
            else:
                # PI sees only subawards for their awards
                cur.execute(
                    """
                    SELECT s.*, a.title as award_title, a.status as award_status
                    FROM subawards s
                    INNER JOIN awards a ON s.award_id = a.award_id
                    WHERE a.created_by_email = %s
                    ORDER BY s.created_at DESC
                    """,
                    (u["email"],)
                )
            subawards_list = cur.fetchall()
            # Convert amounts to float for template rendering
            for sub in subawards_list:
                if sub.get('amount'):
                    sub['amount'] = float(sub['amount'])
            
            # Get available awards for creating new subawards
            cur.execute(
                """
                SELECT award_id, title, amount, status
                FROM awards
                WHERE status = 'Approved'
                ORDER BY title
                """
            )
            awards = cur.fetchall()
            # Convert amounts to float for template rendering
            for award in awards:
                if award.get('amount'):
                    award['amount'] = float(award['amount'])
            awards_map = {a['award_id']: a for a in awards}
            
            cur.close()
        except psycopg2_errors.UndefinedTable:
            # Tables don't exist - just show empty list, no error message
            subawards_list = []
            awards_map = {}
        except Exception as e:
            print(f"DB fetch subawards error: {e}")
            # Return empty list on error
            subawards_list = []
            awards_map = {}
        finally:
            conn.close()
    
    return render_template("subawards.html", subawards=subawards_list, awards_map=awards_map, user=u)


@app.route("/subawards/new")
def subaward_new():
    """Show form to create a new subaward."""
    u = session.get("user")
    if not u:
        return redirect(url_for("home"))
    
    award_id = request.args.get("award_id", type=int)
    
    conn = get_db()
    awards = []
    selected_award = None
    
    if conn is not None:
        try:
            cur = conn.cursor(cursor_factory=RealDictCursor)
            
            if u["role"] == "Admin":
                cur.execute(
                    """
                    SELECT award_id, title, amount, status
                    FROM awards
                    WHERE status = 'Approved'
                    ORDER BY title
                    """
                )
            else:
                cur.execute(
                    """
                    SELECT award_id, title, amount, status
                    FROM awards
                    WHERE status = 'Approved' AND created_by_email = %s
                    ORDER BY title
                    """,
                    (u["email"],)
                )
            awards = cur.fetchall()
            # Convert amounts to float for template rendering
            for award in awards:
                if award.get('amount'):
                    award['amount'] = float(award['amount'])
            
            if award_id:
                # Verify the award exists and user has permission
                if u["role"] == "Admin":
                    cur.execute(
                        "SELECT * FROM awards WHERE award_id = %s",
                        (award_id,)
                    )
                else:
                    cur.execute(
                        "SELECT * FROM awards WHERE award_id = %s AND created_by_email = %s",
                        (award_id, u["email"])
                    )
                selected_award = cur.fetchone()
            
            cur.close()
        except Exception as e:
            print(f"DB fetch awards error: {e}")
        finally:
            conn.close()
    
    return render_template("subaward_new.html", awards=awards, selected_award=selected_award, user=u)


@app.route("/subawards", methods=["POST"])
def subaward_create():
    """Create a new subaward."""
    u = session.get("user")
    if not u:
        return redirect(url_for("home"))
    
    award_id = request.form.get("award_id", type=int)
    subrecipient_name = request.form.get("subrecipient_name", "").strip()
    subrecipient_contact = request.form.get("subrecipient_contact", "").strip()
    subrecipient_email = request.form.get("subrecipient_email", "").strip()
    amount = request.form.get("amount", "").strip()
    start_date = request.form.get("start_date", "").strip()
    end_date = request.form.get("end_date", "").strip()
    description = request.form.get("description", "").strip()
    
    if not award_id or not subrecipient_name or not amount:
        return make_response("Missing required fields", 400)
    
    try:
        amount_val = float(amount)
        if amount_val <= 0:
            return make_response("Amount must be positive", 400)
    except ValueError:
        return make_response("Invalid amount", 400)
    
    conn = get_db()
    if conn is None:
        return make_response("DB connection failed", 500)
    
    try:
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Verify award exists and is approved
        cur.execute(
            """
            SELECT award_id, amount, status, created_by_email
            FROM awards
            WHERE award_id = %s
            """,
            (award_id,)
        )
        award = cur.fetchone()
        
        if not award:
            cur.close()
            conn.close()
            return "Award not found", 404
        
        if award['status'] != 'Approved':
            cur.close()
            conn.close()
            return "Only approved awards can have subawards", 400
        
        # Check permissions
        if u["role"] != "Admin" and award['created_by_email'] != u["email"]:
            cur.close()
            conn.close()
            return "Unauthorized", 403
        
        # Check if subaward amount exceeds award amount
        try:
            cur.execute(
                """
                SELECT COALESCE(SUM(amount), 0) as total_subawards
                FROM subawards
                WHERE award_id = %s AND status != 'Declined'
                """,
                (award_id,)
            )
            result = cur.fetchone()
            total_subawards = float(result['total_subawards'] or 0) if result else 0
        except Exception as table_error:
            # Table might not exist
            print(f"Table access error (subawards might not exist): {table_error}")
            cur.close()
            conn.close()
            return make_response("Database tables not initialized. Please run the schema migration.", 500)
        
        if total_subawards + amount_val > float(award['amount'] or 0):
            cur.close()
            conn.close()
            return make_response("Subaward total exceeds award amount", 400)
        
        # Insert subaward
        cur.execute(
            """
            INSERT INTO subawards (
                award_id, subrecipient_name, subrecipient_contact,
                subrecipient_email, amount, start_date, end_date,
                description, status, created_by_email
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, 'Pending', %s)
            RETURNING subaward_id
            """,
            (
                award_id, subrecipient_name, subrecipient_contact or None,
                subrecipient_email or None, amount_val,
                start_date or None, end_date or None,
                description or None, u["email"]
            )
        )
        subaward_id = cur.fetchone()['subaward_id']
        
        conn.commit()
        cur.close()
        
    except psycopg2_errors.UndefinedTable as table_error:
        print(f"Table does not exist: {table_error}")
        conn.rollback()
        if conn:
            conn.close()
        return redirect(url_for("admin_init_db"))
    except Exception as e:
        print(f"DB create subaward error: {e}")
        import traceback
        traceback.print_exc()
        conn.rollback()
        return make_response(f"Subaward creation failed: {str(e)}", 500)
    finally:
        if conn:
            conn.close()
    
    return redirect(url_for("subaward_view", subaward_id=subaward_id))


@app.route("/subawards/<int:subaward_id>")
def subaward_view(subaward_id):
    """View a subaward details."""
    u = session.get("user")
    if not u:
        return redirect(url_for("home"))
    
    conn = get_db()
    subaward = None
    award = None
    transactions = []
    budget_status = {}
    
    if conn is not None:
        try:
            cur = conn.cursor(cursor_factory=RealDictCursor)
            
            # Get subaward
            try:
                cur.execute(
                    """
                    SELECT s.*, a.title as award_title, a.status as award_status
                    FROM subawards s
                    LEFT JOIN awards a ON s.award_id = a.award_id
                    WHERE s.subaward_id = %s
                    """,
                    (subaward_id,)
                )
                subaward = cur.fetchone()
                # Convert amount to float for template rendering
                if subaward and subaward.get('amount'):
                    subaward['amount'] = float(subaward['amount'])
            except psycopg2_errors.UndefinedTable:
                cur.close()
                conn.close()
                return redirect(url_for("admin_init_db"))
            except Exception as e:
                print(f"Error fetching subaward: {e}")
                cur.close()
                conn.close()
                return f"Error loading subaward: {str(e)}", 500
            
            if not subaward:
                cur.close()
                conn.close()
                return "Subaward not found", 404
            
            # Get parent award first (needed for permission check)
            cur.execute(
                "SELECT * FROM awards WHERE award_id = %s",
                (subaward['award_id'],)
            )
            award = cur.fetchone()
            
            if not award:
                return "Parent award not found", 404
            
            # Check permissions
            if u["role"] != "Admin" and subaward.get('created_by_email') != u["email"]:
                # Also check if user owns the parent award
                if award.get('created_by_email') != u["email"]:
                    return "Unauthorized", 403
            
            # Get transactions (table might not exist, so catch error)
            try:
                cur.execute(
                    """
                    SELECT t.*, u.name as user_name
                    FROM subaward_transactions t
                    LEFT JOIN users u ON t.user_id = u.user_id
                    WHERE t.subaward_id = %s
                    ORDER BY t.date_submitted DESC
                    """,
                    (subaward_id,)
                )
                transactions = cur.fetchall()
                # Convert amounts to float for template rendering
                for txn in transactions:
                    if txn.get('amount'):
                        txn['amount'] = float(txn['amount'])
            except psycopg2_errors.UndefinedTable:
                transactions = []
            
            # Get budget status (table might not exist, so catch error)
            try:
                cur.execute(
                    """
                    SELECT category, allocated_amount, spent_amount, committed_amount
                    FROM subaward_budget_lines
                    WHERE subaward_id = %s
                    """,
                    (subaward_id,)
                )
                budget_lines = cur.fetchall()
            except psycopg2_errors.UndefinedTable:
                budget_lines = []
            
            for line in budget_lines:
                cat = line['category'] or 'Other'
                budget_status[cat] = {
                    'allocated': float(line['allocated_amount'] or 0),
                    'spent': float(line['spent_amount'] or 0),
                    'committed': float(line['committed_amount'] or 0),
                }
            
            # Add transactions to budget
            for txn in transactions:
                cat = txn['category'] or 'Other'
                amount = float(txn['amount'] or 0)
                
                if cat not in budget_status:
                    budget_status[cat] = {'allocated': 0, 'spent': 0, 'committed': 0}
                
                if txn['status'] == 'Approved':
                    budget_status[cat]['spent'] += amount
                elif txn['status'] == 'Pending':
                    budget_status[cat]['committed'] += amount
            
            # Calculate remaining
            for cat in budget_status:
                budget_status[cat]['remaining'] = (
                    budget_status[cat]['allocated'] -
                    budget_status[cat]['spent'] -
                    budget_status[cat]['committed']
                )
            
            cur.close()
        except Exception as e:
            print(f"DB fetch subaward error: {e}")
            import traceback
            traceback.print_exc()
        finally:
            if conn:
                conn.close()
    
    # Calculate totals safely
    totals = {
        'allocated': sum(cat.get('allocated', 0) for cat in budget_status.values()) if budget_status else 0,
        'spent': sum(cat.get('spent', 0) for cat in budget_status.values()) if budget_status else 0,
        'committed': sum(cat.get('committed', 0) for cat in budget_status.values()) if budget_status else 0,
        'remaining': sum(cat.get('remaining', 0) for cat in budget_status.values()) if budget_status else 0,
    }
    
    return render_template(
        "subaward_view.html",
        subaward=subaward,
        award=award,
        transactions=transactions,
        budget_status=budget_status,
        totals=totals,
        user=u
    )


@app.route("/subawards/<int:subaward_id>/approve", methods=["POST"])
def subaward_approve(subaward_id):
    """Approve a subaward (Admin only)."""
    u = session.get("user")
    if not u or u.get("role") != "Admin":
        return redirect(url_for("home"))
    
    conn = get_db()
    if conn is None:
        return make_response("DB connection failed", 500)
    
    try:
        cur = conn.cursor(cursor_factory=RealDictCursor)
        # Check if subaward exists
        cur.execute(
            "SELECT subaward_id FROM subawards WHERE subaward_id = %s",
            (subaward_id,)
        )
        if not cur.fetchone():
            return "Subaward not found", 404
        
        cur.execute(
            "UPDATE subawards SET status = 'Approved' WHERE subaward_id = %s",
            (subaward_id,)
        )
        conn.commit()
        cur.close()
    except Exception as e:
        print(f"DB approve subaward error: {e}")
        conn.rollback()
        return make_response("Approve failed", 500)
    finally:
        conn.close()
    
    return redirect(url_for("subaward_view", subaward_id=subaward_id))


@app.route("/subawards/<int:subaward_id>/delete", methods=["POST"])
def subaward_delete(subaward_id):
    """Delete a subaward."""
    u = session.get("user")
    if not u:
        return redirect(url_for("home"))
    
    conn = get_db()
    if conn is None:
        return make_response("DB connection failed", 500)
    
    try:
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Check permissions
        cur.execute(
            """
            SELECT s.*, a.created_by_email as award_owner
            FROM subawards s
            LEFT JOIN awards a ON s.award_id = a.award_id
            WHERE s.subaward_id = %s
            """,
            (subaward_id,)
        )
        subaward = cur.fetchone()
        
        if not subaward:
            return "Subaward not found", 404
        
        if u["role"] != "Admin" and subaward['created_by_email'] != u["email"] and subaward.get('award_owner') != u["email"]:
            return "Unauthorized", 403
        
        cur.execute("DELETE FROM subawards WHERE subaward_id = %s", (subaward_id,))
        conn.commit()
        cur.close()
        
    except Exception as e:
        print(f"DB delete subaward error: {e}")
        conn.rollback()
        return make_response("Delete failed", 500)
    finally:
        conn.close()
    
    return redirect(url_for("subawards"))


# ========== TRANSACTION SYSTEM ==========

def get_budget_status(award_id):
    """Calculate budget status for an award: allocated, committed, spent, remaining by category."""
    conn = get_db()
    if conn is None:
        return {}
    
    try:
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Get budget lines (allocated amounts by category)
        cur.execute(
            """
            SELECT category, allocated_amount, spent_amount, committed_amount
            FROM budget_lines
            WHERE award_id = %s
            """,
            (award_id,)
        )
        budget_lines = cur.fetchall()
        
        # Get transactions
        cur.execute(
            """
            SELECT category, amount, status
            FROM transactions
            WHERE award_id = %s
            """,
            (award_id,)
        )
        transactions = cur.fetchall()
        
        # Calculate by category
        categories = {}
        pending_by_category = {}
        
        # Initialize from budget_lines (allocated amounts only - transactions are source of truth for spent/committed)
        for line in budget_lines:
            cat = line['category'] or 'Other'
            # Skip "Total" category - we calculate totals separately
            if cat == 'Total':
                continue
            categories[cat] = {
                "allocated": float(line.get("allocated_amount") or 0),
                "spent": 0,  # Will be calculated from transactions
                "committed": 0,  # Will be calculated from transactions
            }
        
        # Calculate spent and committed from transactions (source of truth)
        for txn in transactions:
            cat = txn['category'] or 'Other'
            amount = float(txn['amount'] or 0)
            status = txn['status']
            
            # Map "Other" transactions to "Other Direct Costs" if it exists, otherwise keep as "Other"
            if cat == 'Other' and 'Other Direct Costs' in categories:
                cat = 'Other Direct Costs'
            
            if cat not in categories:
                categories[cat] = {'allocated': 0, 'spent': 0, 'committed': 0}
            
            if status == 'Paid':
                # Paid transactions are actual expenditures (spent)
                categories[cat]['spent'] += amount
            elif status == 'Approved':
                # Approved transactions are committed obligations (not yet paid)
                categories[cat]['committed'] += amount
            elif status == 'Pending':
                # Pending transactions are NOT in committed - they're just requests
                # They should be considered for budget availability check but not stored as committed
                pending_by_category[cat] = pending_by_category.get(cat, 0) + amount
        
        # Calculate remaining
        for cat, vals in categories.items():
            pending_amt = pending_by_category.get(cat, 0.0)
            # Committed = only approved transactions (obligations not yet paid)
            # Spent = only paid transactions (actual expenditures)
            # Pending transactions are NOT in committed, but should be considered for budget availability
            # Remaining = allocated - spent - committed - pending (all reduce available budget for checking)
            vals['remaining'] = max(0, vals['allocated'] - vals['spent'] - vals['committed'] - pending_amt)
        
        cur.close()
        return categories
        
    except Exception as e:
        print(f"Budget status calculation error: {e}")
        return {}
    finally:
        conn.close()


def initialize_budget_lines(award_id):
    """Initialize budget_lines from award's budget breakdown."""
    conn = get_db()
    if conn is None:
        return False
    
    try:
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Get award details
        cur.execute(
            """
            SELECT amount, budget_personnel, budget_equipment, 
                   budget_travel, budget_materials, personnel_json,
                   domestic_travel_json, international_travel_json, materials_json
            FROM awards
            WHERE award_id = %s
            """,
            (award_id,)
        )
        award = cur.fetchone()
        
        if not award:
            cur.close()
            return False
        
        # Calculate budget by category from JSON data
        categories = {}
        total_award = float(award['amount'] or 0)
        def parse_json_field(raw):
            if not raw:
                return []
            if isinstance(raw, (list, dict)):
                return raw
            try:
                return json.loads(raw)
            except (TypeError, json.JSONDecodeError):
                return []
        # Personnel - use rate_per_hour * hours from form
        personnel = parse_json_field(award.get('personnel_json'))
        personnel_total = 0
        personnel_items = []
        for p in personnel:
            if isinstance(p, dict):
                hours_list = p.get('hours', [])
                rate_per_hour = float(p.get('rate_per_hour', 0) or 0)
                total_from_form = float(p.get('total', 0) or 0)
                
                # If total is provided, use it; otherwise calculate from hours * rate
                if total_from_form > 0:
                    personnel_total += total_from_form
                elif isinstance(hours_list, list) and rate_per_hour > 0:
                    total_hours = sum(float(h.get('hours', 0) or 0) for h in hours_list if isinstance(h, dict))
                    personnel_total += total_hours * rate_per_hour
                
                personnel_items.append({
                    'name': p.get('name', 'Unknown'),
                    'position': p.get('position', 'N/A'),
                    'total': total_from_form if total_from_form > 0 else (sum(float(h.get('hours', 0) or 0) for h in hours_list if isinstance(h, dict)) * rate_per_hour if rate_per_hour > 0 else 0)
                })
        categories['Personnel'] = personnel_total
        
        # Travel - use total_amount from new simplified structure
        dom_travel = parse_json_field(award.get('domestic_travel_json'))
        intl_travel = parse_json_field(award.get('international_travel_json'))
        travel_total = 0
        travel_items = []
        
        # Handle new structure (total_amount) and old structure (flight/taxi/food)
        for t in dom_travel + intl_travel:
            if isinstance(t, dict):
                # New structure: total_amount
                total_amount = float(t.get('total_amount', 0) or 0)
                if total_amount > 0:
                    travel_total += total_amount
                    travel_items.append({
                        'description': t.get('description', 'N/A'),
                        'type': 'Domestic' if t in dom_travel else 'International',
                        'amount': total_amount
                    })
                else:
                    # Old structure fallback
                    flight = float(t.get("flight_cost") or t.get("flight") or 0)
                    taxi = float(t.get("taxi_per_day") or t.get("taxi") or 0)
                    food = float(t.get("food_lodge_per_day") or t.get("food_per_day") or t.get("food") or 0)
                    days = float(t.get("days", 0) or 0)
                    old_total = flight + (taxi + food) * days
                    if old_total > 0:
                        travel_total += old_total
                        travel_items.append({
                            'description': t.get('description', 'N/A'),
                            'type': 'Domestic' if t in dom_travel else 'International',
                            'amount': old_total
                        })
        categories["Travel"] = travel_total
        
        # Materials, Equipment, and Other Direct Costs - separate by type
        materials_json = parse_json_field(award.get('materials_json'))
        materials_total = 0
        equipment_total = 0
        other_costs_total = 0
        materials_items = []
        equipment_items = []
        other_costs_items = []
        
        for m in materials_json:
            if isinstance(m, dict):
                cost = float(m.get('cost', 0) or 0)
                item_type = m.get('type', '')
                
                if item_type == 'equipment':
                    equipment_total += cost
                    equipment_items.append({
                        'description': m.get('description', 'N/A'),
                        'amount': cost
                    })
                elif item_type == 'other':
                    other_costs_total += cost
                    other_costs_items.append({
                        'description': m.get('description', 'N/A'),
                        'amount': cost
                    })
                else:
                    # Regular materials
                    materials_total += cost
                    materials_items.append({
                        'description': m.get('description', 'N/A'),
                        'amount': cost
                    })
        
        categories['Materials'] = materials_total
        categories["Equipment"] = equipment_total
        categories['Other Direct Costs'] = other_costs_total
        
        # Other (remaining from total)
        total_allocated = sum(categories.values())
        categories['Other'] = max(0, total_award - total_allocated)
        
        # If no detailed breakdown, allocate everything to "Other"
        if total_allocated == 0 and total_award > 0:
            categories = {"Other": total_award}
        
        # Get subawards total
        cur.execute(
            """
            SELECT COALESCE(SUM(amount), 0) as total_subawards
            FROM subawards
            WHERE award_id = %s AND status != 'Declined'
            """,
            (award_id,)
        )
        subaward_result = cur.fetchone()
        subawards_total = float(subaward_result['total_subawards'] or 0) if subaward_result else 0
        if subawards_total > 0:
            categories['Subawards'] = subawards_total
        
        # Insert/update budget_lines
        for category, amount in categories.items():
            if amount > 0:
                # Check if exists
                cur.execute(
                    """
                    SELECT line_id FROM budget_lines
                    WHERE award_id = %s AND category = %s
                    """,
                    (award_id, category)
                )
                exists = cur.fetchone()
                
                if exists:
                    # Update allocated amount (preserve spent/committed from transactions)
                    # Recalculate remaining based on new allocated amount
                    cur.execute(
                        """
                        UPDATE budget_lines
                        SET allocated_amount = %s
                        WHERE award_id = %s AND category = %s
                        """,
                        (amount, award_id, category)
                    )
                else:
                    # Insert new budget line
                    cur.execute(
                        """
                        INSERT INTO budget_lines (award_id, category, allocated_amount, spent_amount, committed_amount)
                        VALUES (%s, %s, %s, 0, 0)
                        """,
                        (award_id, category, amount)
                    )
        
        
        conn.commit()
        cur.close()
        return True
        
    except Exception as e:
        print(f"Initialize budget lines error: {e}")
        conn.rollback()
        return False
    finally:
        conn.close()


@app.route("/transactions/new")
def transaction_new():
    """Show form to create a new transaction."""
    u = session.get("user")
    if not u:
        return redirect(url_for("home"))
    
    award_id = request.args.get("award_id", type=int)
    if not award_id:
        return redirect(url_for("dashboard"))
    
    # Get award details
    conn = get_db()
    award = None
    if conn is not None:
        try:
            cur = conn.cursor(cursor_factory=RealDictCursor)
            cur.execute(
                """
                SELECT award_id, title, amount, status
                FROM awards
                WHERE award_id = %s AND (created_by_email = %s OR %s = 'Admin')
                """,
                (award_id, u["email"], u["role"])
            )
            award = cur.fetchone()
            cur.close()
        except Exception as e:
            print(f"DB fetch award error: {e}")
        finally:
            conn.close()
    
    if not award:
        return "Award not found", 404
    
    if award['status'] != 'Approved':
        return "Only approved awards can have transactions", 400
    
    # Get budget status
    budget_status = get_budget_status(award_id)
    
    # Get error message from query parameter if present
    error_message = request.args.get("error", "")
    
    return render_template("transaction_new.html", award=award, budget_status=budget_status, user=u, error=error_message)


@app.route("/transactions", methods=["POST"])
def transaction_create():
    """Create a new transaction."""
    u = session.get("user")
    if not u:
        return redirect(url_for("home"))
    
    award_id = request.form.get("award_id", type=int)
    category = request.form.get("category", "").strip()
    description = request.form.get("description", "").strip()
    amount = request.form.get("amount", "").strip()
    date_submitted = request.form.get("date_submitted", "").strip()
    
    if not award_id or not category or not description or not amount or not date_submitted:
        return make_response("Missing required fields", 400)
    
    try:
        amount_val = float(amount)
        if amount_val <= 0:
            return make_response("Amount must be positive", 400)
    except ValueError:
        return make_response("Invalid amount", 400)
    
    conn = get_db()
    if conn is None:
        return make_response("DB connection failed", 500)
    
    try:
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Verify award exists and is approved
        cur.execute(
            """
            SELECT award_id, status, created_by_email
            FROM awards
            WHERE award_id = %s
            """,
            (award_id,)
        )
        award = cur.fetchone()
        
        if not award:
            return "Award not found", 404
        
        if award['status'] != 'Approved':
            return "Only approved awards can have transactions", 400
        
        # Check permissions
        if u["role"] != "Admin" and award['created_by_email'] != u["email"]:
            return "Unauthorized", 403
        
        # Get user_id
        cur.execute("SELECT user_id FROM users WHERE email = %s", (u["email"],))
        user_row = cur.fetchone()
        user_id = user_row['user_id'] if user_row else None
        
        # Map "Other" to "Other Direct Costs" if that category exists in budget
        # Check what categories exist in budget_lines
        cur.execute(
            """
            SELECT category FROM budget_lines
            WHERE award_id = %s
            """,
            (award_id,)
        )
        existing_categories = [row['category'] for row in cur.fetchall()]
        
        # If transaction is "Other" but "Other Direct Costs" exists, map it
        transaction_category = category
        if category == 'Other' and 'Other Direct Costs' in existing_categories:
            transaction_category = 'Other Direct Costs'
        
        # Check budget availability (only if budget has been allocated)
        budget_status = get_budget_status(award_id)
        cat_budget = budget_status.get(transaction_category, {})
        allocated = cat_budget.get('allocated', 0)
        remaining = cat_budget.get('remaining', 0)
        
        # Only check budget if there's an allocated amount for this category
        # If no budget allocated yet, allow the transaction (it will create the budget line)
        # Remaining = allocated - spent - committed, so check against remaining
        if allocated > 0 and amount_val > remaining:
            # Redirect back to form with error message (URL encode the error)
            error_msg = f"Insufficient budget for {transaction_category}. Remaining: ${remaining:,.2f}, Requested: ${amount_val:,.2f}"
            return redirect(url_for('transaction_new', award_id=award_id, error=error_msg))
        
        # Insert transaction
        cur.execute(
            """
            INSERT INTO transactions (award_id, user_id, category, description, amount, date_submitted, status)
            VALUES (%s, %s, %s, %s, %s, %s, 'Pending')
            RETURNING transaction_id
            """,
            (award_id, user_id, transaction_category, description, amount_val, date_submitted)
        )
        transaction_id = cur.fetchone()['transaction_id']
        
        # Pending transactions should NOT be added to committed_amount
        # They will be added to committed only when approved
        # Budget check is done above, so we just insert the transaction
        
        conn.commit()
        cur.close()
        
    except Exception as e:
        print(f"DB create transaction error: {e}")
        import traceback
        traceback.print_exc()
        conn.rollback()
        return make_response(f"Transaction creation failed: {str(e)}", 500)
    finally:
        conn.close()
    
    return redirect(url_for("transactions_list", award_id=award_id))


@app.route("/transactions")
def transactions_list():
    """List all transactions for an award or user."""
    u = session.get("user")
    if not u:
        return redirect(url_for("home"))
    
    award_id = request.args.get("award_id", type=int)
    
    conn = get_db()
    transactions = []
    award = None
    
    if conn is not None:
        try:
            cur = conn.cursor(cursor_factory=RealDictCursor)
            
            if award_id:
                # Get transactions for specific award
                cur.execute(
                    """
                    SELECT t.*, a.title as award_title, u.name as user_name
                    FROM transactions t
                    LEFT JOIN awards a ON t.award_id = a.award_id
                    LEFT JOIN users u ON t.user_id = u.user_id
                    WHERE t.award_id = %s
                    ORDER BY t.date_submitted DESC, t.transaction_id DESC
                    """,
                    (award_id,)
                )
                transactions = cur.fetchall()
                
                # Get award details
                cur.execute("SELECT * FROM awards WHERE award_id = %s", (award_id,))
                award = cur.fetchone()
            else:
                # Get all transactions for user
                if u["role"] == "Admin":
                    cur.execute(
                        """
                        SELECT t.*, a.title as award_title, u.name as user_name
                        FROM transactions t
                        LEFT JOIN awards a ON t.award_id = a.award_id
                        LEFT JOIN users u ON t.user_id = u.user_id
                        ORDER BY t.date_submitted DESC, t.transaction_id DESC
                        """
                    )
                else:
                    cur.execute(
                        """
                        SELECT t.*, a.title as award_title, u.name as user_name
                        FROM transactions t
                        LEFT JOIN awards a ON t.award_id = a.award_id
                        LEFT JOIN users u ON t.user_id = u.user_id
                        WHERE a.created_by_email = %s
                        ORDER BY t.date_submitted DESC, t.transaction_id DESC
                        """,
                        (u["email"],)
                    )
                transactions = cur.fetchall()
            
            cur.close()
        except Exception as e:
            print(f"DB fetch transactions error: {e}")
        finally:
            conn.close()
    
    return render_template("transactions_list.html", transactions=transactions, award=award, user=u)


@app.route("/awards/<int:award_id>/budget")
def budget_status(award_id):
    """Show budget status dashboard for an award."""
    u = session.get("user")
    if not u:
        return redirect(url_for("home"))
    
    conn = get_db()
    award = None
    budget_status_data = {}
    
    if conn is not None:
        try:
            cur = conn.cursor(cursor_factory=RealDictCursor)
            cur.execute(
                """
                SELECT * FROM awards
                WHERE award_id = %s AND (created_by_email = %s OR %s = 'Admin')
                """,
                (award_id, u["email"], u["role"])
            )
            award = cur.fetchone()
            cur.close()
        except Exception as e:
            print(f"DB fetch award error: {e}")
        finally:
            conn.close()
    
    if not award:
        return "Award not found", 404
    
    # Initialize/recalculate budget lines if award is approved
    # This ensures budget always matches the current form data
    if award['status'] == 'Approved':
        # Always recalculate from form data - it will update allocated amounts
        initialize_budget_lines(award_id)
    
    budget_status_data = get_budget_status(award_id)
    
    # ---- SUBAWARD COMMITTED AMOUNT (future obligations, not yet paid) ----
    conn = get_db()
    subaward_committed = 0.0

    if conn is not None:
        try:
            cur = conn.cursor(cursor_factory=RealDictCursor)
            cur.execute(
                """
                SELECT COALESCE(SUM(amount), 0) AS subaward_total
                FROM subawards
                WHERE award_id = %s
                  AND status = 'Approved'
                """,
                (award_id,)
            )
            row = cur.fetchone()
            subaward_committed = float(row["subaward_total"]) if row else 0.0
            cur.close()
        except Exception as e:
            print(f"DB fetch subaward total error: {e}")
        finally:
            conn.close()

    # Helper function to parse JSON fields
    def parse_json_field(raw):
        if not raw:
            return []
        if isinstance(raw, (list, dict)):
            return raw
        try:
            return json.loads(raw)
        except (TypeError, json.JSONDecodeError):
            return []
    
    # Get detailed items for each category
    conn = get_db()
    personnel_items = []
    travel_items = []
    materials_items = []
    equipment_items = []
    other_costs_items = []
    subawards_list = []
    
    if conn is not None:
        try:
            cur = conn.cursor(cursor_factory=RealDictCursor)
            
            # Personnel items
            personnel_json = parse_json_field(award.get('personnel_json'))
            for p in personnel_json:
                if isinstance(p, dict):
                    hours_list = p.get('hours', [])
                    rate_per_hour = float(p.get('rate_per_hour', 0) or 0)
                    total_from_form = float(p.get('total', 0) or 0)
                    
                    if total_from_form > 0:
                        total = total_from_form
                    elif isinstance(hours_list, list) and rate_per_hour > 0:
                        total_hours = sum(float(h.get('hours', 0) or 0) for h in hours_list if isinstance(h, dict))
                        total = total_hours * rate_per_hour
                    else:
                        total = 0
                    
                    if total > 0:
                        personnel_items.append({
                            'name': p.get('name', 'Unknown'),
                            'position': p.get('position', 'N/A'),
                            'amount': total
                        })
            
            # Travel items
            dom_travel = parse_json_field(award.get('domestic_travel_json'))
            intl_travel = parse_json_field(award.get('international_travel_json'))
            for t in dom_travel + intl_travel:
                if isinstance(t, dict):
                    total_amount = float(t.get('total_amount', 0) or 0)
                    if total_amount > 0:
                        travel_items.append({
                            'description': t.get('description', 'N/A'),
                            'type': 'Domestic' if t in dom_travel else 'International',
                            'amount': total_amount
                        })
            
            # Materials, Equipment, Other Direct Costs
            materials_json = parse_json_field(award.get('materials_json'))
            for m in materials_json:
                if isinstance(m, dict):
                    cost = float(m.get('cost', 0) or 0)
                    item_type = m.get('type', '')
                    
                    if item_type == 'equipment':
                        equipment_items.append({
                            'description': m.get('description', 'N/A'),
                            'amount': cost
                        })
                    elif item_type == 'other':
                        other_costs_items.append({
                            'description': m.get('description', 'N/A'),
                            'amount': cost
                        })
                    else:
                        materials_items.append({
                            'description': m.get('description', 'N/A'),
                            'amount': cost
                        })
            
            # Subawards
            cur.execute(
                """
                SELECT subaward_id, subrecipient_name, amount, status
                FROM subawards
                WHERE award_id = %s AND status != 'Declined'
                ORDER BY created_at DESC
                """,
                (award_id,)
            )
            subawards_list = cur.fetchall()
            for sub in subawards_list:
                if sub.get('amount'):
                    sub['amount'] = float(sub['amount'])
            
            cur.close()
        except Exception as e:
            print(f"DB fetch items error: {e}")
        finally:
            conn.close()
    
    # Calculate totals (ensure all are floats)
    # Spent = only paid expenses (approved transactions)
    # Committed = only future obligations (pending transactions + approved subawards)
    allocated = float(sum(cat.get('allocated', 0) for cat in budget_status_data.values()))
    spent = float(sum(cat.get('spent', 0) for cat in budget_status_data.values()))
    committed_tx = float(sum(cat.get('committed', 0) for cat in budget_status_data.values()))  # Pending transactions only
    
    # Include approved subawards as committed (future obligations, not yet paid)
    committed = committed_tx + subaward_committed
    
    # Remaining = allocated - spent - committed
    remaining = allocated - spent - committed
    
    totals = {
        'allocated': allocated,
        'spent': spent,
        'committed': committed,
        'remaining': remaining,
    }
    
    # Convert award.amount to float and calculate remaining budget
    # Remaining = Total Award - Spent - Committed (both reduce available budget)
    total_award_amount = float(award.get('amount') or 0)
    remaining_budget = total_award_amount - totals['spent'] - totals['committed']
    
    # Ensure award.amount is float for template display
    if award.get('amount'):
        award['amount'] = float(award['amount'])
    
    u = session.get("user")
    return render_template(
        "budget_status.html",
        award=award,
        budget_status=budget_status_data,
        totals=totals,
        total_award_amount=total_award_amount,
        remaining_budget=remaining_budget,
        personnel_items=personnel_items,
        travel_items=travel_items,
        materials_items=materials_items,
        equipment_items=equipment_items,
        other_costs_items=other_costs_items,
        subawards_list=subawards_list,
        user=u or {}
    )


@app.route("/transactions/<int:transaction_id>/approve", methods=["POST"])
def transaction_approve(transaction_id):
    """Approve a transaction (Admin/Finance only).
    
    This moves the transaction from Pending to Approved status.
    Approved transactions are in COMMITTED (not spent) until payment is processed.
    Policy compliance is checked before approval.
    """
    u = session.get("user")
    if not u or u.get("role") not in ("Admin", "Finance"):
        return redirect(url_for("home"))
    
    conn = get_db()
    if conn is None:
        return make_response("DB connection failed", 500)
    
    try:
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Get transaction details with award info
        cur.execute(
            """
            SELECT t.*, a.status as award_status, a.title, a.sponsor_type, a.amount as award_amount,
                   a.start_date, a.end_date
            FROM transactions t
            JOIN awards a ON t.award_id = a.award_id
            WHERE t.transaction_id = %s
            """,
            (transaction_id,)
        )
        txn = cur.fetchone()
        
        if not txn:
            return "Transaction not found", 404
        
        if txn['status'] != 'Pending':
            return "Transaction already processed", 400
        
        if txn['award_status'] != 'Approved':
            return "Award must be approved", 400
        
        # Check policy compliance before approval
        award_data = {
            'title': txn.get('title', ''),
            'sponsor_type': txn.get('sponsor_type', ''),
            'amount': txn.get('award_amount', 0),
            'start_date': txn.get('start_date', ''),
            'end_date': txn.get('end_date', '')
        }
        compliance_results = check_transaction_compliance(txn, award_data)
        
        # Store compliance results as JSON
        compliance_json = json.dumps(compliance_results)
        
        # Check if any policy is non-compliant
        has_non_compliant = any(
            result.get('result') == 'non-compliant' 
            for result in compliance_results.values() 
            if isinstance(result, dict)
        )
        
        # Update transaction status to Approved and store compliance notes
        cur.execute(
            """
            UPDATE transactions 
            SET status = 'Approved', compliance_notes = %s
            WHERE transaction_id = %s
            """,
            (compliance_json, transaction_id)
        )
        
        # Map transaction category to budget category
        txn_category = txn['category'] or 'Other'
        # Check if "Other Direct Costs" exists in budget_lines
        cur.execute(
            """
            SELECT category FROM budget_lines
            WHERE award_id = %s
            """,
            (txn['award_id'],)
        )
        existing_categories = [row['category'] for row in cur.fetchall()]
        if txn_category == 'Other' and 'Other Direct Costs' in existing_categories:
            txn_category = 'Other Direct Costs'
        
        # When approving, add the transaction to committed_amount
        # Pending transactions are NOT in committed, only Approved transactions are
        # Check if budget line exists
        cur.execute(
            """
            SELECT line_id FROM budget_lines
            WHERE award_id = %s AND category = %s
            """,
            (txn['award_id'], txn_category)
        )
        budget_line = cur.fetchone()
        
        if budget_line:
            cur.execute(
                """
                UPDATE budget_lines
                SET committed_amount = committed_amount + %s
                WHERE award_id = %s AND category = %s
                """,
                (txn['amount'], txn['award_id'], txn_category)
            )
        else:
            # Create budget line if it doesn't exist
            cur.execute(
                """
                INSERT INTO budget_lines (award_id, category, allocated_amount, spent_amount, committed_amount)
                VALUES (%s, %s, 0, 0, %s)
                """,
                (txn['award_id'], txn_category, txn['amount'])
            )
        
        conn.commit()
        cur.close()
        
        # If non-compliant, warn but still allow approval (admin decision)
        if has_non_compliant:
            print(f"WARNING: Transaction {transaction_id} approved despite non-compliant policy check")
        
    except Exception as e:
        print(f"DB approve transaction error: {e}")
        import traceback
        traceback.print_exc()
        conn.rollback()
        return make_response("Approve failed", 500)
    finally:
        conn.close()
    
    return redirect(url_for("transactions_list", award_id=txn['award_id']))


@app.route("/transactions/<int:transaction_id>/pay", methods=["POST"])
def transaction_pay(transaction_id):
    """Process payment for an approved transaction (Admin/Finance only).
    
    This moves the transaction from Approved to Paid status.
    The amount moves from committed to spent in the budget.
    """
    u = session.get("user")
    if not u or u.get("role") not in ("Admin", "Finance"):
        return redirect(url_for("home"))
    
    conn = get_db()
    if conn is None:
        return make_response("DB connection failed", 500)
    
    try:
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Get transaction details
        cur.execute(
            """
            SELECT t.transaction_id, t.award_id, t.category, t.amount, t.status, a.status as award_status
            FROM transactions t
            JOIN awards a ON t.award_id = a.award_id
            WHERE t.transaction_id = %s
            """,
            (transaction_id,)
        )
        txn = cur.fetchone()
        
        if not txn:
            cur.close()
            conn.close()
            return "Transaction not found", 404
        
        if txn['status'] != 'Approved':
            cur.close()
            conn.close()
            return "Only approved transactions can be paid", 400
        
        if txn['award_status'] != 'Approved':
            cur.close()
            conn.close()
            return "Award must be approved", 400
        
        # Convert amount to float
        amount_val = float(txn['amount'] or 0)
        award_id = txn['award_id']
        
        # Update transaction status to Paid
        cur.execute(
            "UPDATE transactions SET status = 'Paid' WHERE transaction_id = %s",
            (transaction_id,)
        )
        
        # Map transaction category to budget category
        txn_category = txn['category'] or 'Other'
        # Check if "Other Direct Costs" exists in budget_lines
        cur.execute(
            """
            SELECT category FROM budget_lines
            WHERE award_id = %s
            """,
            (award_id,)
        )
        existing_categories = [row['category'] for row in cur.fetchall()]
        if txn_category == 'Other' and 'Other Direct Costs' in existing_categories:
            txn_category = 'Other Direct Costs'
        
        # Move from committed to spent in budget_lines
        # First ensure budget line exists
        cur.execute(
            """
            SELECT line_id FROM budget_lines
            WHERE award_id = %s AND category = %s
            """,
            (award_id, txn_category)
        )
        budget_line = cur.fetchone()
        
        if budget_line:
            cur.execute(
                """
                UPDATE budget_lines
                SET committed_amount = GREATEST(0, committed_amount - %s),
                    spent_amount = spent_amount + %s
                WHERE award_id = %s AND category = %s
                """,
                (amount_val, amount_val, award_id, txn_category)
            )
        else:
            # Create budget line if it doesn't exist
            cur.execute(
                """
                INSERT INTO budget_lines (award_id, category, allocated_amount, spent_amount, committed_amount)
                VALUES (%s, %s, 0, %s, 0)
                """,
                (award_id, txn_category, amount_val)
            )
        
        conn.commit()
        cur.close()
        
    except Exception as e:
        print(f"DB pay transaction error: {e}")
        import traceback
        traceback.print_exc()
        if conn:
            conn.rollback()
        return make_response(f"Payment processing failed: {str(e)}", 500)
    finally:
        if conn:
            conn.close()
    
    # Get award_id for redirect (use the one we extracted earlier)
    return redirect(url_for("transactions_list", award_id=award_id))


@app.route("/transactions/<int:transaction_id>/decline", methods=["POST"])
def transaction_decline(transaction_id):
    """Decline a transaction (Admin/Finance only)."""
    u = session.get("user")
    if not u or u.get("role") not in ("Admin", "Finance"):
        return redirect(url_for("home"))
    
    conn = get_db()
    if conn is None:
        return make_response("DB connection failed", 500)
    
    try:
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Get transaction details
        cur.execute(
            "SELECT * FROM transactions WHERE transaction_id = %s",
            (transaction_id,)
        )
        txn = cur.fetchone()
        
        if not txn:
            return "Transaction not found", 404
        
        # Can decline either Pending or Approved transactions (but not Paid)
        if txn['status'] not in ('Pending', 'Approved'):
            return "Transaction already processed or paid", 400
        
        award_id = txn['award_id']
        old_status = txn['status']
        
        # Update transaction status
        cur.execute(
            "UPDATE transactions SET status = 'Declined' WHERE transaction_id = %s",
            (transaction_id,)
        )
        
        # Map transaction category to budget category
        txn_category = txn['category'] or 'Other'
        # Check if "Other Direct Costs" exists in budget_lines
        cur.execute(
            """
            SELECT category FROM budget_lines
            WHERE award_id = %s
            """,
            (award_id,)
        )
        existing_categories = [row['category'] for row in cur.fetchall()]
        if txn_category == 'Other' and 'Other Direct Costs' in existing_categories:
            txn_category = 'Other Direct Costs'
        
        # Only remove from committed if the transaction was Approved
        # Pending transactions are NOT in committed, so nothing to remove
        if old_status == 'Approved':
            cur.execute(
                """
                SELECT line_id FROM budget_lines
                WHERE award_id = %s AND category = %s
                """,
                (award_id, txn_category)
            )
            budget_line = cur.fetchone()
            
            if budget_line:
                cur.execute(
                    """
                    UPDATE budget_lines
                    SET committed_amount = GREATEST(0, committed_amount - %s)
                    WHERE award_id = %s AND category = %s
                    """,
                    (txn['amount'], award_id, txn_category)
                )
        
        conn.commit()
        cur.close()
        
    except Exception as e:
        print(f"DB decline transaction error: {e}")
        conn.rollback()
        return make_response("Decline failed", 500)
    finally:
        conn.close()
    
    return redirect(url_for("transactions_list", award_id=award_id))


@app.route("/settings")
def settings():
    u = session.get("user")
    if not u:
        return redirect(url_for("home"))
    return render_template("settings.html", user=u)


@app.route("/profile")
def profile():
    u = session.get("user")
    if not u:
        return redirect(url_for("home"))
    awards = []
    conn = get_db()
    if conn is not None:
        try:
            cur = conn.cursor(cursor_factory=RealDictCursor)
            if u["role"] == "Admin":
                # Admin: see all awards
                cur.execute(
                    """
                    SELECT award_id, title, sponsor_type, amount, start_date, end_date, status, created_at
                    FROM awards
                    ORDER BY created_at DESC
                    """
                )
            else:
                # PI: only their own awards
                cur.execute(
                    """
                    SELECT award_id, title, sponsor_type, amount, start_date, end_date, status, created_at
                    FROM awards
                    WHERE created_by_email=%s
                    ORDER BY created_at DESC
                    """,
                    (u["email"],),
                )
            awards = cur.fetchall()
            cur.close()
        except Exception as e:
            print(f"DB fetch awards (profile) error: {e}")
        finally:
            conn.close()
    stats = {
        "total_awards": len(awards),
        "active_awards": sum(
            1 for a in awards
            if (a.get("status") or "").lower() == "approved"
        ),
        "latest_award": awards[0] if awards else None,
    }

    return render_template("profile.html", user=u, awards=awards, stats=stats)


def parse_policy_file(filepath):
    """Parse a policy text file and extract key summary points."""
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            content = f.read()
    except Exception as e:
        print(f"Error reading policy file {filepath}: {e}")
        return None
    
    lines = content.split('\n')
    
    # Extract title (first non-empty line)
    title = ""
    description = ""
    key_points = []
    seen_thresholds = set()
    
    i = 0
    while i < len(lines):
        line = lines[i].strip()
        
        # Skip empty lines at the start
        if not line:
            i += 1
            continue
        
        # Get title (first non-empty line)
        if not title:
            title = line
            i += 1
            continue
        
        # Skip description lines (lines in parentheses or descriptive text)
        if line.startswith('(') or (i < 5 and not line[0].isdigit()):
            i += 1
            continue
        
        # Extract key thresholds - look for lines with dollar amounts and thresholds
        if "$" in line:
            line_lower = line.lower()
            # Personnel threshold
            if ("personnel" in line_lower or "salary" in line_lower) and "12,000" in line and "personnel" not in seen_thresholds:
                key_points.append("Personnel: $12,000 per person per year (covers salaries, wages, graduate assistants)")
                seen_thresholds.add("personnel")
            # Travel threshold
            elif "travel" in line_lower and "5,000" in line and "travel" not in seen_thresholds:
                key_points.append("Travel: $5,000 per trip (conferences, fieldwork, research meetings)")
                seen_thresholds.add("travel")
            # Equipment threshold
            elif "equipment" in line_lower and "8,000" in line and "equipment" not in seen_thresholds:
                key_points.append("Equipment: $8,000 per item (research equipment, computers, lab instruments)")
                seen_thresholds.add("equipment")
            # Materials threshold
            elif ("materials" in line_lower or "supplies" in line_lower) and "5,000" in line and "materials" not in seen_thresholds:
                key_points.append("Materials & Supplies: $5,000 per item (chemicals, reagents, consumables)")
                seen_thresholds.add("materials")
            # Other Direct Costs threshold
            elif "other direct" in line_lower and "5,000" in line and "other" not in seen_thresholds:
                key_points.append("Other Direct Costs: $5,000 per item (publication fees, participant incentives)")
                seen_thresholds.add("other")
        
        # Stop at POLICY HIERARCHY section
        if line.startswith("POLICY HIERARCHY"):
            break
        
        i += 1
    
    # Add general policy summary based on policy type
    if "FEDERAL" in title.upper():
        key_points.append("Not allowed: First-class travel, personal expenses, or charging for work not performed")
        key_points.append("Required: Economy-class travel only; Fly America Act applies for international travel")
        key_points.append("All expenses must be allowable, allocable, reasonable, and consistently treated")
    elif "SPONSOR" in title.upper():
        key_points.insert(0, "Must follow both Federal and sponsor-specific requirements")
        key_points.append("Not allowed: Personnel or equipment not listed in approved proposal")
        key_points.append("Changes require sponsor approval before spending")
    elif "UNIVERSITY" in title.upper():
        key_points.insert(0, "Must follow Federal and Sponsor requirements")
        key_points.append("Not allowed: Administrative staff unless 100% dedicated to project")
        key_points.append("All expenses must directly support research objectives")
    
    return {
        "title": title,
        "description": "",  # Don't show description
        "key_points": key_points
    }


@app.route("/policies/university")
def university_policies():
    """Display all policies from the policy files."""
    u = session.get("user")
    
    # Parse all three policy files
    base_dir = os.path.dirname(os.path.abspath(__file__))
    
    policy_files = [
        {
            "file": os.path.join(base_dir, "policies", "federal_policy.txt"),
            "level": "Federal",
            "order": 1
        },
        {
            "file": os.path.join(base_dir, "policies", "sponsor_policy.txt"),
            "level": "Sponsor",
            "order": 2
        },
        {
            "file": os.path.join(base_dir, "policies", "university_policy.txt"),
            "level": "University",
            "order": 3
        }
    ]
    
    policy_data = []
    for pf in policy_files:
        parsed = parse_policy_file(pf["file"])
        if parsed:
            # Clean up sponsor policy title
            title = parsed["title"]
            if "SPONSOR POLICY" in title.upper():
                title = "SPONSOR POLICY"
            
            policy_data.append({
                "level": pf["level"],
                "title": title,
                "description": parsed["description"],
                "key_points": parsed["key_points"],
                "order": pf["order"]
            })
    
    # Sort by order
    policy_data.sort(key=lambda x: x["order"])
    
    return render_template("policies_university.html", policies=policy_data, user=u or {})


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("home"))


# ========== LLM POLICY COMPLIANCE CHECKING ==========

def read_policy_file(policy_name):
    """Read policy text from file."""
    policy_path = os.path.join("policies", f"{policy_name}_policy.txt")
    try:
        with open(policy_path, "r", encoding="utf-8") as f:
            return f.read()
    except FileNotFoundError:
        print(f"Policy file not found: {policy_path}")
        return ""
    except Exception as e:
        print(f"Error reading policy file {policy_path}: {e}")
        return ""


def format_award_for_llm(award, personnel, domestic_travel, international_travel, materials, equipment=None, other_direct=None):
    """Format award data into a structured text for LLM analysis."""
    
    # Calculate personnel budget from personnel data
    personnel_budget = 0.0
    if personnel:
        for p in personnel:
            if isinstance(p, dict):
                # Use total if provided, otherwise calculate from hours * rate
                total_from_form = float(p.get('total', 0) or 0)
                if total_from_form > 0:
                    personnel_budget += total_from_form
                else:
                    hours_list = p.get('hours', [])
                    rate_per_hour = float(p.get('rate_per_hour', 0) or 0)
                    if isinstance(hours_list, list) and rate_per_hour > 0:
                        total_hours = sum(float(h.get('hours', 0) or 0) for h in hours_list if isinstance(h, dict))
                        personnel_budget += total_hours * rate_per_hour
    
    # Calculate travel budget
    travel_budget = 0.0
    if domestic_travel:
        for t in domestic_travel:
            if isinstance(t, dict):
                total_amount = float(t.get('total_amount', 0) or 0)
                if total_amount > 0:
                    travel_budget += total_amount
    if international_travel:
        for t in international_travel:
            if isinstance(t, dict):
                total_amount = float(t.get('total_amount', 0) or 0)
                if total_amount > 0:
                    travel_budget += total_amount
    
    # Calculate equipment budget from equipment_json
    equipment_budget = 0.0
    if equipment:
        for e in equipment:
            if isinstance(e, dict):
                cost = float(e.get('cost', 0) or 0)
                equipment_budget += cost
    
    # Calculate materials budget from materials_json
    materials_budget = 0.0
    if materials:
        for m in materials:
            if isinstance(m, dict):
                cost = float(m.get('cost', 0) or 0)
                materials_budget += cost
    
    # Calculate other direct costs budget
    other_direct_budget = 0.0
    if other_direct:
        for o in other_direct:
            if isinstance(o, dict):
                cost = float(o.get('cost', 0) or 0)
                other_direct_budget += cost
    
    award_text = f"""
AWARD INFORMATION:
- Title: {award.get('title', 'N/A')}
- Sponsor Type: {award.get('sponsor_type', 'N/A')}
- Total Amount: ${float(award.get('amount', 0) or 0):,.2f}
- Start Date: {award.get('start_date', 'N/A')}
- End Date: {award.get('end_date', 'N/A')}
- Department: {award.get('department', 'N/A')}
- College: {award.get('college', 'N/A')}

BUDGET BREAKDOWN (TOTALS - for reference only):
- Personnel Budget: ${personnel_budget:,.2f}
- Equipment Budget: ${equipment_budget:,.2f}
- Travel Budget: ${travel_budget:,.2f}
- Materials Budget: ${materials_budget:,.2f}
- Other Direct Costs Budget: ${other_direct_budget:,.2f}

IMPORTANT: The budget totals above are for reference. Policy compliance is checked PER ITEM, not by total budget.
Equipment items are checked individually against $8,000 per item.
Materials items are checked individually against $5,000 per item.
Travel is checked per trip against $5,000 per trip.
"""
    
    if personnel:
        award_text += "\nPERSONNEL DETAILS:\n"
        for p in personnel:
            if isinstance(p, dict):
                name = p.get('name', 'Unknown')
                role = p.get('position', 'N/A') or p.get('role', 'N/A')
                hours_list = p.get('hours', [])
                rate_per_hour = float(p.get('rate_per_hour', 0) or 0)
                total_from_form = float(p.get('total', 0) or 0)
                
                total_hours = 0
                if isinstance(hours_list, list):
                    for h in hours_list:
                        if isinstance(h, dict):
                            total_hours += float(h.get('hours', 0) or 0)
                
                # Calculate total if not provided
                if total_from_form > 0:
                    personnel_total = total_from_form
                elif total_hours > 0 and rate_per_hour > 0:
                    personnel_total = total_hours * rate_per_hour
                else:
                    personnel_total = 0
                
                award_text += f"- {name} ({role}): {total_hours} hours"
                if rate_per_hour > 0:
                    award_text += f", Rate: ${rate_per_hour:,.2f}/hour"
                if personnel_total > 0:
                    award_text += f", Total: ${personnel_total:,.2f}"
                award_text += "\n"
    
    if domestic_travel:
        award_text += "\nDOMESTIC TRAVEL:\n"
        for t in domestic_travel:
            if isinstance(t, dict):
                description = t.get('description', 'N/A')
                total_amount = float(t.get('total_amount', 0) or 0)
                if total_amount > 0:
                    award_text += f"- {description}: Total Estimated Amount: ${total_amount:,.2f}\n"
                else:
                    # Fallback to old format if total_amount not available
                    flight = t.get('flight', 0) or t.get('flight_cost', 0)
                    taxi = t.get('taxi', 0) or t.get('taxi_per_day', 0)
                    food = t.get('food', 0) or t.get('food_per_day', 0)
                    days = t.get('days', 0) or t.get('num_days', 0)
                    award_text += f"- {description}: Flight: ${float(flight or 0):,.2f}, Taxi/day: ${float(taxi or 0):,.2f}, Food/day: ${float(food or 0):,.2f}, Days: {days}\n"
    
    if international_travel:
        award_text += "\nINTERNATIONAL TRAVEL:\n"
        for t in international_travel:
            if isinstance(t, dict):
                description = t.get('description', 'N/A')
                total_amount = float(t.get('total_amount', 0) or 0)
                if total_amount > 0:
                    award_text += f"- {description}: Total Estimated Amount: ${total_amount:,.2f}\n"
                else:
                    # Fallback to old format if total_amount not available
                    flight = t.get('flight', 0) or t.get('flight_cost', 0)
                    taxi = t.get('taxi', 0) or t.get('taxi_per_day', 0)
                    food = t.get('food', 0) or t.get('food_per_day', 0)
                    days = t.get('days', 0) or t.get('num_days', 0)
                    award_text += f"- {description}: Flight: ${float(flight or 0):,.2f}, Taxi/day: ${float(taxi or 0):,.2f}, Food/day: ${float(food or 0):,.2f}, Days: {days}\n"
    
    if equipment:
        award_text += "\nEQUIPMENT (Section 2 - $8,000 per item threshold):\n"
        for e in equipment:
            if isinstance(e, dict):
                description = e.get('description', 'N/A')
                cost = float(e.get('cost', 0) or 0)
                award_text += f"- Item: {description}, Cost: ${cost:,.2f} (check if this INDIVIDUAL item exceeds $8,000)\n"
        award_text += "NOTE: Check each equipment item INDIVIDUALLY against the $8,000 per item threshold. Do NOT sum all equipment items.\n"
    
    if materials:
        award_text += "\nMATERIALS & SUPPLIES (Section 4 - $5,000 per item threshold):\n"
        for m in materials:
            if isinstance(m, dict):
                description = m.get('description', 'N/A')
                cost = float(m.get('cost', 0) or 0)
                award_text += f"- Item: {description}, Cost: ${cost:,.2f} (check if this INDIVIDUAL item exceeds $5,000)\n"
        award_text += "NOTE: Check each materials item INDIVIDUALLY against the $5,000 per item threshold. Do NOT sum all materials items.\n"
    
    if other_direct:
        award_text += "\nOTHER DIRECT COSTS:\n"
        for o in other_direct:
            if isinstance(o, dict):
                description = o.get('description', 'N/A')
                cost = float(o.get('cost', 0) or 0)
                award_text += f"- {description}, Cost: ${cost:,.2f}\n"
    
    return award_text


def check_policy_compliance(award, personnel, domestic_travel, international_travel, materials, equipment=None, other_direct=None):
    """
    Check award compliance against University, Sponsor, and Federal policies using LLM.
    Returns a dict with compliance results for each policy level.
    """
    # Read policy files
    university_policy = read_policy_file("university")
    federal_policy = read_policy_file("federal")
    sponsor_policy = read_policy_file("sponsor")
    
    # Format award data
    award_text = format_award_for_llm(award, personnel, domestic_travel, international_travel, materials, equipment, other_direct)
    
    # Get OpenAI API key from environment
    api_key = os.getenv("OPENAI_API_KEY")
    if not api_key:
        return {
            "error": "OpenAI API key not configured",
            "university": {"result": "unknown", "reason": "API key missing"},
            "federal": {"result": "unknown", "reason": "API key missing"},
            "sponsor": {"result": "unknown", "reason": "API key missing"}
        }
    
    # Initialize OpenAI client
    client = OpenAI(api_key=api_key)
    
    results = {}
    
    # Check each policy level
    policy_checks = [
        ("university", "University", university_policy),
        ("federal", "Federal", federal_policy),
        ("sponsor", "Sponsor", sponsor_policy)
    ]
    
    for key, name, policy_text in policy_checks:
        if not policy_text:
            results[key] = {"result": "unknown", "reason": f"{name} policy text not available"}
            continue
        
        try:
            # Determine priority level for context
            priority_note = ""
            if name == "Federal":
                priority_note = "NOTE: Federal policy has HIGHEST PRIORITY. Any violation must result in 'non-compliant'."
            elif name == "Sponsor":
                priority_note = "NOTE: Sponsor policy must follow Federal requirements. Check both Federal and Sponsor rules."
            elif name == "University":
                priority_note = "NOTE: University policy is lowest priority but must still be followed. Check if it conflicts with Federal/Sponsor rules."
            
            prompt = f"""You are an AI Policy Compliance Officer for a Post-Award Research Budget Management System.

Your job is to check whether a research award complies with {name} policy.

CRITICAL: You must base every decision ONLY on the policy text provided. Do not assume or invent any rules.

{priority_note}

═══════════════════════════════════════════════════════════════════════════════
CRITICAL RULE: THRESHOLDS ARE PER ITEM, NOT TOTAL
═══════════════════════════════════════════════════════════════════════════════

Equipment and Materials & Supplies are COMPLETELY SEPARATE categories with DIFFERENT thresholds:
- Equipment has a threshold of $8,000 PER ITEM (Section 2 in policies)
- Materials & Supplies has a threshold of $5,000 PER ITEM (Section 4 in policies)
- These categories must be evaluated INDEPENDENTLY. Do NOT combine their costs.
- An equipment item over $8,000 violates Equipment policy, NOT Materials policy.
- A materials item over $5,000 violates Materials policy, NOT Equipment policy.
- IMPORTANT: The threshold is PER ITEM, NOT the total. Each individual item must be checked separately.
- For example: If you have 3 materials items costing $2,000, $1,500, and $1,800, they are ALL compliant because each is under $5,000 per item.
- The TOTAL of all items in a category does NOT need to be under the threshold - only each INDIVIDUAL item.
- Do NOT add up all equipment items and check against $8,000 - check each equipment item individually.
- Do NOT add up all materials items and check against $5,000 - check each materials item individually.

EXAMPLE OF CORRECT EVALUATION:
- If Materials section shows: Item A = $2,000, Item B = $1,500, Item C = $1,800
  → ALL THREE ITEMS ARE COMPLIANT (each is under $5,000 per item)
  → The total ($5,300) does NOT matter - only individual items matter
  
- If Materials section shows: Item A = $2,000, Item B = $6,000
  → Item A is COMPLIANT ($2,000 < $5,000)
  → Item B is NON-COMPLIANT ($6,000 > $5,000 per item threshold)
  → Report Item B as the violation, NOT the total

WRONG: "Materials total $8,500 exceeds $5,000 threshold" ❌
RIGHT: "Each materials item is checked individually. Item X costs $6,200 which exceeds $5,000 per item threshold" ✅

═══════════════════════════════════════════════════════════════════════════════

POLICY TEXT:
{policy_text}

AWARD DATA:
{award_text}

Analyze the award against the {name} policy above. Provide a comprehensive, human-like assessment that:
1. Explains WHY the award is compliant or non-compliant based on policy requirements
2. References specific policy sections and requirements
3. For compliant items: Explain that it follows policy guidelines, meets requirements, and does not violate any rules
4. For non-compliant items: Clearly state what policy is violated and why
5. Avoid simply stating budget thresholds (e.g., "below $5000") - instead explain policy compliance in context
6. CRITICALLY: Evaluate Equipment and Materials & Supplies as SEPARATE categories with their respective thresholds ($8,000 for Equipment, $5,000 for Materials & Supplies)
7. MOST IMPORTANT: Check each item INDIVIDUALLY. The threshold is PER ITEM, NOT the total. For example:
   - If you have 5 materials items costing $1,000 each, they are ALL compliant (each is under $5,000)
   - If you have 1 materials item costing $6,000, it is NON-COMPLIANT (exceeds $5,000 per item)
   - Do NOT add up all items in a category and check the total - check EACH item separately
   - Equipment total and Materials total are separate - do NOT combine them

Your output must be a JSON object in this exact format:
{{
  "result": "compliant" | "non-compliant" | "unknown",
  "reason": "Comprehensive explanation that reads naturally, explains policy compliance, references specific policy sections, and explains why the award follows or violates policy requirements. Write as if explaining to a colleague, not just listing thresholds."
}}

Example of good explanation for compliant: "This award complies with {name} policy requirements. The personnel expenses are within acceptable limits per person per year as specified in Section 1, and all listed personnel directly contribute to project aims. Travel expenses follow policy guidelines for research-related travel and do not exceed per-trip thresholds. Each equipment item is individually checked and all are under the $8,000 per item threshold. Each materials item is individually checked and all are under the $5,000 per item threshold. The award follows all applicable rules and does not violate any policy restrictions."

Example of good explanation for non-compliant: "This award violates {name} policy in Section 2 (Equipment). One equipment item (High-Performance Workstation) costs $9,500, which exceeds the $8,000 per item threshold without prior approval as required by policy. Note: This is checked per item, not by total equipment budget. Additionally, one materials item (Specialized Reagents) costs $6,200, which exceeds the $5,000 per item threshold in Section 4. These violations must be addressed before approval."

Only return the JSON object, nothing else."""

            response = client.chat.completions.create(
                model="gpt-4o-mini",  # Using gpt-4o-mini for cost efficiency
                messages=[
                    {"role": "system", "content": "You are a policy compliance officer. Provide comprehensive, human-like explanations that explain policy compliance in context. Always respond with valid JSON only."},
                    {"role": "user", "content": prompt}
                ],
                temperature=0.4,  # Slightly higher for more natural language
                max_tokens=600  # Increased for more comprehensive explanations
            )
            
            # Parse JSON response
            response_text = response.choices[0].message.content.strip()
            # Remove markdown code blocks if present
            if response_text.startswith("```"):
                response_text = response_text.split("```")[1]
                if response_text.startswith("json"):
                    response_text = response_text[4:]
                response_text = response_text.strip()
            
            compliance_result = json.loads(response_text)
            results[key] = compliance_result
            
        except json.JSONDecodeError as e:
            print(f"Error parsing JSON response for {name} policy: {e}")
            response_text_str = response_text if 'response_text' in locals() else "No response received"
            print(f"Response was: {response_text_str}")
            results[key] = {"result": "unknown", "reason": f"Error parsing LLM response: {str(e)}"}
        except Exception as e:
            print(f"Error checking {name} policy compliance: {e}")
            import traceback
            traceback.print_exc()
            results[key] = {"result": "unknown", "reason": f"Error: {str(e)}"}
    
    return results


def check_transaction_compliance(transaction, award):
    """
    Check transaction compliance against University, Sponsor, and Federal policies using LLM.
    Returns a dict with compliance results for each policy level.
    """
    # Read policy files
    university_policy = read_policy_file("university")
    federal_policy = read_policy_file("federal")
    sponsor_policy = read_policy_file("sponsor")
    
    # Format transaction data for LLM
    transaction_text = f"""
Transaction Details:
- Category: {transaction.get('category', 'N/A')}
- Description: {transaction.get('description', 'N/A')}
- Amount: ${transaction.get('amount', 0):,.2f}
- Date Submitted: {transaction.get('date_submitted', 'N/A')}

Award Context:
- Title: {award.get('title', 'N/A')}
- Sponsor: {award.get('sponsor_type', 'N/A')}
- Amount: ${award.get('amount', 0):,.2f}
- Start Date: {award.get('start_date', 'N/A')}
- End Date: {award.get('end_date', 'N/A')}
"""
    
    # Get OpenAI API key from environment
    api_key = os.getenv("OPENAI_API_KEY")
    if not api_key:
        return {
            "error": "OpenAI API key not configured",
            "university": {"result": "unknown", "reason": "API key missing"},
            "federal": {"result": "unknown", "reason": "API key missing"},
            "sponsor": {"result": "unknown", "reason": "API key missing"}
        }
    
    # Initialize OpenAI client
    client = OpenAI(api_key=api_key)
    
    results = {}
    
    # Check each policy level
    policy_checks = [
        ("university", "University", university_policy),
        ("federal", "Federal", federal_policy),
        ("sponsor", "Sponsor", sponsor_policy)
    ]
    
    for key, name, policy_text in policy_checks:
        if not policy_text:
            results[key] = {"result": "unknown", "reason": f"{name} policy text not available"}
            continue
        
        try:
            # Determine priority level for context
            priority_note = ""
            if name == "Federal":
                priority_note = "NOTE: Federal policy has HIGHEST PRIORITY. Any violation must result in 'non-compliant'."
            elif name == "Sponsor":
                priority_note = "NOTE: Sponsor policy must follow Federal requirements. Check both Federal and Sponsor rules."
            elif name == "University":
                priority_note = "NOTE: University policy is lowest priority but must still be followed. Check if it conflicts with Federal/Sponsor rules."
            
            prompt = f"""You are an AI Policy Compliance Officer for a Post-Award Research Budget Management System.

Your job is to check whether a TRANSACTION (spending request) complies with {name} policy.

CRITICAL: You must base every decision ONLY on the policy text provided. Do not assume or invent any rules.

{priority_note}

IMPORTANT TRANSACTION POLICY RULES:
- Once a transaction is approved, it is recorded as a committed cost and no longer counted as "requested."
- Committed costs do NOT increase the spent total until the university actually pays the expense.
- The remaining balance is immediately reduced by the committed amount to prevent overspending.
- All approved transactions must undergo a compliance verification before payment (allowable, allocable, reasonable).
- Payments cannot exceed the approved amount unless a new approval request is submitted.
- The transaction must follow procurement or travel rules based on category (equipment, supplies, airfare, etc.).
- All post-approval activities must comply with Federal, Sponsor, and University policies.

POLICY TEXT:
{policy_text}

TRANSACTION DATA:
{transaction_text}

Analyze the transaction against the {name} policy above. Provide a comprehensive, human-like assessment that:
1. Explains WHY the transaction is compliant or non-compliant based on policy requirements
2. References specific policy sections and requirements
3. For compliant items: Explain that it follows policy guidelines, is allowable and reasonable, and does not violate any rules
4. For non-compliant items: Clearly state what policy is violated and why
5. Avoid simply stating budget thresholds (e.g., "below $5000") - instead explain policy compliance in context
6. Consider whether the transaction is necessary for the research, properly categorized, and follows procurement/travel rules

Your output must be a JSON object in this exact format:
{{
  "result": "compliant" | "non-compliant" | "unknown",
  "reason": "Comprehensive explanation that reads naturally, explains policy compliance, references specific policy sections, and explains why the transaction follows or violates policy requirements. Write as if explaining to a colleague, not just listing thresholds."
}}

Example of good explanation for compliant: "This transaction complies with {name} policy requirements. The {transaction.get('category', 'expense')} expense is necessary for the research project as described, falls within acceptable policy limits, and follows the procurement guidelines specified in Section 2. The amount is reasonable and allocable to the award, and the transaction does not violate any policy restrictions. It is properly categorized and meets all applicable policy requirements."

Example of good explanation for non-compliant: "This transaction violates {name} policy in Section 3 (Travel). The transaction exceeds the $5,000 per-trip threshold without prior approval as required by policy. Additionally, the description suggests personal travel expenses which are explicitly prohibited. These violations must be addressed before approval."

Only return the JSON object, nothing else."""

            response = client.chat.completions.create(
                model="gpt-4o-mini",
                messages=[
                    {"role": "system", "content": "You are a policy compliance officer. Provide comprehensive, human-like explanations that explain policy compliance in context. Always respond with valid JSON only."},
                    {"role": "user", "content": prompt}
                ],
                temperature=0.4,  # Slightly higher for more natural language
                max_tokens=600  # Increased for more comprehensive explanations
            )
            
            # Parse JSON response
            response_text = response.choices[0].message.content.strip()
            # Remove markdown code blocks if present
            if response_text.startswith("```"):
                response_text = response_text.split("```")[1]
                if response_text.startswith("json"):
                    response_text = response_text[4:]
                response_text = response_text.strip()
            
            compliance_result = json.loads(response_text)
            results[key] = compliance_result
            
        except json.JSONDecodeError as e:
            print(f"Error parsing JSON response for {name} policy: {e}")
            response_text_str = response_text if 'response_text' in locals() else "No response received"
            print(f"Response was: {response_text_str}")
            results[key] = {"result": "unknown", "reason": f"Error parsing LLM response: {str(e)}"}
        except Exception as e:
            print(f"Error checking {name} policy compliance: {e}")
            import traceback
            traceback.print_exc()
            results[key] = {"result": "unknown", "reason": f"Error: {str(e)}"}
    
    return results


@app.route("/awards/<int:award_id>/check-compliance", methods=["POST"])
def check_award_compliance(award_id):
    """Check policy compliance for an award using LLM."""
    u = session.get("user")
    if not u:
        return make_response(json.dumps({"error": "Not authenticated"}), 401, {"Content-Type": "application/json"})
    
    # Only Admin can check compliance
    if u.get("role") != "Admin":
        return make_response(json.dumps({"error": "Unauthorized"}), 403, {"Content-Type": "application/json"})
    
    conn = get_db()
    if conn is None:
        return make_response(json.dumps({"error": "DB connection failed"}), 500, {"Content-Type": "application/json"})
    
    try:
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Get award
        cur.execute("SELECT * FROM awards WHERE award_id=%s", (award_id,))
        award = cur.fetchone()
        
        if not award:
            cur.close()
            conn.close()
            return make_response(json.dumps({"error": "Award not found"}), 404, {"Content-Type": "application/json"})
        
        # Parse JSON fields
        def parse_json(field_name):
            raw = award.get(field_name)
            if raw is None:
                return []
            if isinstance(raw, (dict, list)):
                return raw
            try:
                return json.loads(raw)
            except Exception:
                return []
        
        personnel = parse_json("personnel_json")
        domestic_travel = parse_json("domestic_travel_json")
        international_travel = parse_json("international_travel_json")
        materials = parse_json("materials_json")
        equipment = parse_json("equipment_json")
        other_direct = parse_json("other_direct_json")
        
        cur.close()
        conn.close()
        
        # Check compliance
        compliance_results = check_policy_compliance(award, personnel, domestic_travel, international_travel, materials, equipment, other_direct)
        
        # Store results in database (optional - update ai_review_notes)
        if "error" not in compliance_results:
            conn = get_db()
            if conn:
                try:
                    cur = conn.cursor()
                    notes = json.dumps(compliance_results)
                    cur.execute(
                        "UPDATE awards SET ai_review_notes = %s WHERE award_id = %s",
                        (notes, award_id)
                    )
                    conn.commit()
                    cur.close()
                except Exception as e:
                    print(f"Error saving compliance results: {e}")
                finally:
                    conn.close()
        
        return make_response(json.dumps(compliance_results, indent=2), 200, {"Content-Type": "application/json"})
        
    except Exception as e:
        print(f"Error checking compliance: {e}")
        return make_response(json.dumps({"error": str(e)}), 500, {"Content-Type": "application/json"})


@app.route("/transactions/<int:transaction_id>/check-compliance", methods=["POST"])
def check_transaction_compliance_route(transaction_id):
    """Check policy compliance for a transaction using LLM."""
    u = session.get("user")
    if not u:
        return make_response(json.dumps({"error": "Not authenticated"}), 401, {"Content-Type": "application/json"})
    
    # Only Admin/Finance can check compliance
    if u.get("role") not in ("Admin", "Finance"):
        return make_response(json.dumps({"error": "Unauthorized"}), 403, {"Content-Type": "application/json"})
    
    conn = get_db()
    if conn is None:
        return make_response(json.dumps({"error": "DB connection failed"}), 500, {"Content-Type": "application/json"})
    
    try:
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Get transaction with award info
        cur.execute(
            """
            SELECT t.*, a.title, a.sponsor_type, a.amount as award_amount,
                   a.start_date, a.end_date
            FROM transactions t
            JOIN awards a ON t.award_id = a.award_id
            WHERE t.transaction_id = %s
            """,
            (transaction_id,)
        )
        txn = cur.fetchone()
        
        if not txn:
            cur.close()
            conn.close()
            return make_response(json.dumps({"error": "Transaction not found"}), 404, {"Content-Type": "application/json"})
        
        # Format award data for compliance check
        award_data = {
            'title': txn.get('title', ''),
            'sponsor_type': txn.get('sponsor_type', ''),
            'amount': float(txn.get('award_amount', 0) or 0),
            'start_date': txn.get('start_date', ''),
            'end_date': txn.get('end_date', '')
        }
        
        cur.close()
        conn.close()
        
        # Check compliance
        compliance_results = check_transaction_compliance(txn, award_data)
        
        # Store results in database (update compliance_notes)
        if "error" not in compliance_results:
            conn = get_db()
            if conn:
                try:
                    cur = conn.cursor()
                    notes = json.dumps(compliance_results)
                    cur.execute(
                        "UPDATE transactions SET compliance_notes = %s WHERE transaction_id = %s",
                        (notes, transaction_id)
                    )
                    conn.commit()
                    cur.close()
                except Exception as e:
                    print(f"Error saving transaction compliance results: {e}")
                finally:
                    conn.close()
        
        return make_response(json.dumps(compliance_results, indent=2), 200, {"Content-Type": "application/json"})
        
    except Exception as e:
        print(f"Error checking transaction compliance: {e}")
        import traceback
        traceback.print_exc()
        return make_response(json.dumps({"error": str(e)}), 500, {"Content-Type": "application/json"})


@app.route("/admin/init-db", methods=["GET", "POST"])
def admin_init_db():
    """Initialize database and update constraints."""
    # Update transactions status constraint
    update_transactions_status_constraint()
    """Admin route to manually initialize database schema."""
    u = session.get("user")
    if not u or u.get("role") != "Admin":
        return "Unauthorized - Admin access required", 403
    
    if request.method == "POST":
        conn = get_db()
        if conn is None:
            return "Database connection failed", 500
        
        try:
            cur = conn.cursor()
            schema_file = os.path.join(os.path.dirname(__file__), "schema_postgresql.sql")
            if os.path.exists(schema_file):
                with open(schema_file, 'r') as f:
                    schema_sql = f.read()
                
                # Remove comments and execute
                lines = schema_sql.split('\n')
                cleaned_lines = []
                for line in lines:
                    stripped = line.strip()
                    if stripped and not stripped.startswith('--'):
                        cleaned_lines.append(line)
                cleaned_sql = '\n'.join(cleaned_lines)
                
                try:
                    cur.execute(cleaned_sql)
                    conn.commit()
                    message = "✓ Database schema initialized successfully!"
                except Exception as full_error:
                    # Try statement by statement
                    statements = schema_sql.split(';')
                    executed = 0
                    for statement in statements:
                        statement = statement.strip()
                        if statement and not statement.startswith('--') and not statement.upper().startswith('SELECT'):
                            try:
                                cur.execute(statement)
                                executed += 1
                            except Exception as stmt_error:
                                error_msg = str(stmt_error)
                                if 'already exists' not in error_msg.lower():
                                    print(f"Statement error: {stmt_error}")
                    conn.commit()
                    message = f"✓ Database schema initialized! ({executed} statements executed)"
                
                cur.close()
                conn.close()
                # Redirect to subawards page after successful initialization
                return redirect(url_for("subawards"))
            else:
                return f"Schema file not found: {schema_file}", 404
        except Exception as e:
            import traceback
            traceback.print_exc()
            return f"Error initializing database: {str(e)}", 500
    
    # GET request - automatically initialize and redirect
    # No confirmation page, just do it
    conn = get_db()
    if conn is None:
        return "Database connection failed", 500
    
    try:
        cur = conn.cursor()
        schema_file = os.path.join(os.path.dirname(__file__), "schema_postgresql.sql")
        if os.path.exists(schema_file):
            with open(schema_file, 'r') as f:
                schema_sql = f.read()
            
            # Execute statements
            statements = schema_sql.split(';')
            for statement in statements:
                statement = statement.strip()
                if statement and not statement.startswith('--') and not statement.upper().startswith('SELECT'):
                    try:
                        cur.execute(statement)
                    except Exception:
                        pass  # Ignore errors (tables might already exist)
            conn.commit()
        cur.close()
        conn.close()
    except Exception:
        pass
    
    # Always redirect to subawards
    return redirect(url_for("subawards"))


if __name__ == "__main__":
    init_db_if_needed()
    # Ensure transactions constraint includes 'Paid' status
    update_transactions_status_constraint()
    app.run(debug=True, port=8000)
